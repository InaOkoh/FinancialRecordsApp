import os
import json
from datetime import datetime
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from tkinter import simpledialog
from tkinter import filedialog
import customtkinter as ctk

ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

try:
    from tkcalendar import DateEntry
except ImportError:
    DateEntry = None

# Attempt to import optional/external packages (as specified in requirements.txt)
try:
    import openpyxl
    import openpyxl.styles
    import openpyxl.utils
except ImportError:
    openpyxl = None

try:
    import matplotlib
    import matplotlib.pyplot as plt
except ImportError:
    matplotlib = None
    plt = None

try:
    import fpdf
except ImportError:
    fpdf = None

class CheckboxDropdown(ctk.CTkFrame):
    def __init__(self, master, values, command=None, width=200, **kwargs):
        super().__init__(master, fg_color="transparent", **kwargs)
        self.values = values
        self.command = command
        self.selected_categories = ["All Categories"]
        self.vars = {}
        
        self.label = ctk.CTkLabel(self, text="Filter by Financial Report Category:", font=("Segoe UI", 11), text_color="gray")
        self.label.pack(side="top", anchor="w", pady=(0, 2))
        
        self.button = ctk.CTkButton(self, text="Select Categories ▼", width=width, 
                                    border_width=1,
                                    command=self.toggle_dropdown)
        self.button.pack(fill="both", expand=True)
        self.toplevel = None

    def get(self):
        return self.selected_categories
        
    def toggle_dropdown(self):
        if self.toplevel and self.toplevel.winfo_exists():
            self.toplevel.destroy()
            self.toplevel = None
        else:
            self.toplevel = ctk.CTkToplevel(self.winfo_toplevel())
            self.toplevel.overrideredirect(True)
            self.toplevel.attributes("-topmost", True)
            
            x = self.button.winfo_rootx()
            y = self.button.winfo_rooty() + self.button.winfo_height()
            self.toplevel.geometry(f"{self.button.winfo_width()}x200+{x}+{y}")
            
            frame = ctk.CTkFrame(self.toplevel, border_width=1)
            frame.pack(fill="both", expand=True)
            
            btn_apply = ctk.CTkButton(frame, text="Apply Filter", height=24, command=self.apply)
            btn_apply.pack(side="bottom", pady=5, padx=5, fill="x")

            sf = ctk.CTkScrollableFrame(frame, fg_color="transparent")
            sf.pack(side="top", fill="both", expand=True, padx=2, pady=2)
            
            for val in self.values:
                var = ctk.StringVar(value="1" if val in self.selected_categories else "0")
                self.vars[val] = var
                cb = ctk.CTkCheckBox(sf, text=val, variable=var, onvalue="1", offvalue="0",
                                     command=lambda v=val: self.on_checkbox_click(v))
                cb.pack(padx=5, pady=5, anchor="w")

    def on_checkbox_click(self, toggled_val):
        if toggled_val == "All Categories" and self.vars["All Categories"].get() == "1":
            for val, var in self.vars.items():
                if val != "All Categories":
                    var.set("0")
        elif toggled_val != "All Categories" and self.vars[toggled_val].get() == "1":
            if "All Categories" in self.vars:
                self.vars["All Categories"].set("0")

    def apply(self):
        self.selected_categories = [val for val, var in self.vars.items() if var.get() == "1"]
        if not self.selected_categories or "All Categories" in self.selected_categories:
            self.selected_categories = ["All Categories"]
            
        if self.command:
            self.command()
            
        self.toggle_dropdown()


def number_to_words(amount):
    """
    Converts a numeric amount into English words representation.
    """
    try:
        if isinstance(amount, str):
            amount = amount.replace(",", "")
        val = float(amount)
    except (ValueError, TypeError):
        return ""
    
    if val < 0:
        return "Negative " + number_to_words(abs(val))
        
    integer_part = int(val)
    fractional_part = int(round((val - integer_part) * 100))
    
    ones = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine",
            "Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen",
            "Seventeen", "Eighteen", "Nineteen"]
    tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]
    thousands = ["", "Thousand", "Million", "Billion"]
    
    def convert_group(n):
        h = n // 100
        t = (n % 100) // 10
        o = n % 10
        
        words = []
        if h > 0:
            words.append(ones[h] + " Hundred")
        
        if t >= 2:
            if o > 0:
                words.append(tens[t] + "-" + ones[o])
            else:
                words.append(tens[t])
        elif t == 1 or o > 0:
            words.append(ones[t * 10 + o])
            
        return " ".join(words)
        
    if integer_part == 0:
        int_str = "Zero"
    else:
        groups = []
        temp = integer_part
        while temp > 0:
            groups.append(temp % 1000)
            temp = temp // 1000
            
        group_words = []
        for i, grp in enumerate(groups):
            if grp > 0:
                grp_word = convert_group(grp)
                if thousands[i]:
                    grp_word += " " + thousands[i]
                group_words.append(grp_word)
                
        group_words.reverse()
        int_str = " ".join(group_words)
        
    if integer_part == 1:
        result = int_str + " Naira"
    else:
        result = int_str + " Naira"
        
    if fractional_part > 0:
        cents_str = convert_group(fractional_part)
        if fractional_part == 1:
            result += " and " + cents_str + " Kobo"
        else:
            result += " and " + cents_str + " Kobo"
    else:
        result += " Only"
    return result

def make_combo_spinner(parent, default_val, vals):
    frame = ctk.CTkFrame(parent, fg_color="transparent")
    combo = ctk.CTkComboBox(frame, values=vals, font=("Segoe UI", 13), width=90)
    combo.set(str(default_val))
    combo.pack(side="left", padx=(0, 2))
    
    btn_frame = ctk.CTkFrame(frame, fg_color="transparent")
    btn_frame.pack(side="left")
    
    frame._command = None
    def update_vals(curr_val):
        new_vals = [str(y) for y in range(curr_val, curr_val + 10)]
        combo.configure(values=new_vals)

    def fire_cmd():
        if frame._command:
            frame._command(combo.get())
            
    def on_select(val):
        try:
            update_vals(int(val))
        except: pass
        fire_cmd()
        
    combo.configure(command=on_select)
    
    def add():
        try:
            val = int(combo.get()) + 1
            combo.set(str(val))
            update_vals(val)
            fire_cmd()
        except: pass
        
    def sub():
        try:
            val = int(combo.get()) - 1
            combo.set(str(val))
            update_vals(val)
            fire_cmd()
        except: pass
        
    ctk.CTkButton(btn_frame, text="▲", width=20, height=14, font=("Segoe UI", 10), command=add).pack(side="top", pady=(0, 1))
    ctk.CTkButton(btn_frame, text="▼", width=20, height=14, font=("Segoe UI", 10), command=sub).pack(side="bottom")
    
    def config(**kwargs):
        if "command" in kwargs:
            frame._command = kwargs.pop("command")
        if kwargs:
            combo.configure(**kwargs)
    
    def custom_set(val):
        combo.set(val)
        try:
            update_vals(int(val))
        except: pass

    frame.get = combo.get
    frame.set = custom_set
    frame.configure = config
    frame.delete = lambda first, last=None: custom_set("")
    frame.insert = lambda index, text: custom_set(text)
    return frame

class VoucherPDF(fpdf.FPDF):
    def __init__(self, company_info=None, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.company_info = company_info or {}

    def header(self):
        import os
        # Document Header
        c_name = self.company_info.get("name", "")
        c_addr = self.company_info.get("address", "")
        c_phone = self.company_info.get("phone", "")
        c_email = self.company_info.get("email", "")
        c_web = self.company_info.get("website", "")
        c_logo = self.company_info.get("logo", "")

        if c_logo and os.path.isfile(c_logo):
            try:
                # Top right corner for portrait (width=210).
                self.image(c_logo, x=175, y=5, w=30)
            except Exception:
                pass

        if c_name:
            self.set_font("Helvetica", "B", 16)
            self.set_text_color(30, 58, 138)
            self.set_x(55)
            self.multi_cell(100, 8, c_name, align="C")

            self.set_font("Helvetica", "", 10)
            self.set_text_color(75, 85, 99)
            if c_addr:
                self.set_x(55)
                self.multi_cell(100, 6, c_addr, align="C")
            if c_phone:
                self.set_x(55)
                self.multi_cell(100, 6, f"Tel: {c_phone}", align="C")
            if c_email:
                self.set_x(55)
                self.multi_cell(100, 6, f"Email: {c_email}", align="C")
            if c_web:
                self.set_x(55)
                self.multi_cell(100, 6, c_web, align="C")
            self.ln(5)

        self.set_font("Helvetica", "B", 18)
        self.set_text_color(30, 58, 138)  # Primary dark blue
        self.cell(0, 10, "Financial Records System", new_x="LMARGIN", new_y="NEXT", align="L")

        self.set_font("Helvetica", "", 10)
        self.set_text_color(75, 85, 99)   # Muted gray
        self.cell(0, 5, "Official Transaction Voucher", new_x="LMARGIN", new_y="NEXT", align="L")

        # Decorative line
        self.set_draw_color(13, 148, 136)  # Accent Teal
        self.set_line_width(0.8)
        y_line = self.get_y()
        self.line(20, y_line, 190, y_line)
        self.ln(8)
        
    def footer(self):
        self.set_y(-20)
        self.set_font("Helvetica", "I", 8)
        self.set_text_color(156, 163, 175) # Light gray
        self.set_draw_color(229, 231, 235)
        self.set_line_width(0.3)
        self.line(20, self.get_y() - 2, 190, self.get_y() - 2)
        
        self.cell(0, 10, "This is a computer-generated document and does not require a physical signature unless specified.", align="L")
        self.set_x(170)
        self.cell(20, 10, f"Page {self.page_no()}", align="R")

class App(ctk.CTk):
    """
    Main application window class for the Financial Records Manager.
    Inherits from customtkinter.CTk to create the primary window.
    """
    def __init__(self):
        super().__init__()

        # Initialize configuration
        self.init_config()

        # Define modern flat theme colors as tuples (Light, Dark)
        self.theme_colors = {
            "bg": ("#F8F9FA", "#0F172A"),
            "card": ("#FFFFFF", "#1E293B"),
            "text": ("#1A2530", "#F1F5F9"),
            "accent": ("#2B5FDF", "#3B82F6"),
            "green": ("#10B981", "#34D399"),
            "red": ("#EF4444", "#F87171"),
            "border": ("#E2E8F0", "#334155"),
            "text_muted": ("#64748B", "#94A3B8")
        }

        # Configure CustomTkinter default themes
        theme = self.config_data.get("theme", "light")
        ctk.set_appearance_mode(theme)
        ctk.set_default_color_theme("blue")

        # Initialize theme variable
        self.theme_var = tk.StringVar(value=theme)

        # Initialize workbook tracker
        self.current_workbook_path = None

        # Configure window properties
        self.is_workbook_valid = True
        self.update_app_title()
        window_width = 1200
        window_height = 800
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        center_x = int((screen_width - window_width) / 2)
        center_y = int((screen_height - window_height) / 2)
        self.geometry(f"{window_width}x{window_height}+{center_x}+{center_y}")
        self.minsize(800, 600)

        # Apply premium styles and colors
        self.configure_styles()

        # Build standard UI components
        # Top menu bar removed as per user request.
        self.create_main_layout()

        # Disable inputs at startup since no file is active
        self.set_forms_state("disabled")

    def configure_styles(self):
        """
        Defines a custom, clean styling system using ttk styles.
        """
        self.apply_theme()

    def switch_theme(self, theme_name):
        """Switches the theme, saves preference to config, and updates widgets in real time."""
        self.config_data["theme"] = theme_name
        self.theme_var.set(theme_name)
        self.save_config()
        ctk.set_appearance_mode(theme_name)
        self.apply_theme()

    def update_widget_colors(self, widget):
        """Recursively updates non-ttk standard Tkinter widgets to the new theme colors."""
        widget_class = widget.winfo_class()
        is_ctk = any(cls.__name__.startswith("CTk") for cls in type(widget).__mro__)

        # Standard Tk Widgets
        try:
            if not is_ctk:
                if widget_class == "Label":
                    if widget == getattr(self, "status_label", None):
                        widget.configure(bg=self.colors["bg"])
                        if not self.current_workbook_path:
                            widget.configure(fg=self.colors["text_muted"])
                        else:
                            curr_fg = widget.cget("fg")
                            if curr_fg != "#dc2626":
                                widget.configure(fg=self.colors["accent"])
                    elif hasattr(widget, "master") and type(widget.master).__name__.startswith("CTk") and type(widget.master).__name__ not in ("CTkFrame", "CTkScrollableFrame", "CTkTabview", "CTk"):
                        pass
                    else:
                        widget.configure(bg=self.colors["card"], fg=self.colors["text"])
                elif widget_class == "Frame":
                    # Listbox container frame has Listbox inside it
                    has_listbox = False
                    for c in widget.winfo_children():
                        if c.winfo_class() == "Listbox":
                            has_listbox = True
                            break
                    if has_listbox:
                        widget.configure(bg=self.colors["border"])
                    else:
                        widget.configure(bg=self.colors["bg"])
                elif widget_class == "Listbox":
                    widget.configure(
                        bg=self.colors["card"], 
                        fg=self.colors["text"],
                        selectbackground=self.colors["primary"],
                        selectforeground="#ffffff"
                    )
                elif widget_class == "Toplevel":
                    widget.configure(bg=self.colors["bg"])
                elif widget_class == "Text":
                    state = widget.cget("state")
                    bg = self.colors["bg"] if state == "disabled" else self.colors["card"]
                    widget.configure(
                        bg=bg,
                        fg=self.colors["text"],
                        insertbackground=self.colors["text"]
                    )
        except tk.TclError:
            pass
            
        # Recurse for all children
        for child in widget.winfo_children():
            self.update_widget_colors(child)

    def apply_theme(self):
        """Applies light or dark theme colors to ttk styles and active widgets."""
        theme = self.config_data.get("theme", "light")

        if theme == "dark":
            self.colors = {
                "bg": "#0F172A",
                "card": "#1E293B",
                "primary": "#3B82F6",
                "accent": "#3B82F6",
                "text": "#F1F5F9",
                "text_muted": "#94A3B8",
                "border": "#334155",
                "green": "#34D399",
                "red": "#F87171"
            }
        else:
            self.colors = {
                "bg": "#F8F9FA",
                "card": "#FFFFFF",
                "primary": "#2B5FDF",
                "accent": "#2B5FDF",
                "text": "#1A2530",
                "text_muted": "#64748B",
                "border": "#E2E8F0",
                "green": "#10B981",
                "red": "#EF4444"
            }

        # Update root background
        self.configure(fg_color=self.colors["bg"])

        self.style = ttk.Style()
        if "clam" in self.style.theme_names():
            self.style.theme_use("clam")

        # Configure standard layouts
        self.style.configure("TNotebook", background=self.colors["bg"], borderwidth=0)
        self.style.configure("TNotebook.Tab",
                              background=self.colors["border"],
                              foreground=self.colors["text_muted"],
                              padding=[20, 8],
                              font=("Segoe UI", 13, "bold"),
                              borderwidth=0)

        self.style.map("TNotebook.Tab",
                       background=[("selected", self.colors["primary"]), ("active", self.colors["accent"])],
                       foreground=[("selected", "#ffffff"), ("active", "#ffffff")])

        self.style.configure("TFrame", background=self.colors["bg"])
        self.style.configure("Card.TFrame", background=self.colors["card"], relief="flat", borderwidth=0)

        self.style.configure("TLabel", background=self.colors["bg"], foreground=self.colors["text"], font=("Segoe UI", 13))
        self.style.configure("Header.TLabel", background=self.colors["card"], foreground=self.colors["primary"], font=("Segoe UI", 18, "bold"))
        self.style.configure("Subheader.TLabel", background=self.colors["card"], foreground=self.colors["text_muted"], font=("Segoe UI", 13, "italic"))
        self.style.configure("CardLabel.TLabel", background=self.colors["card"], foreground=self.colors["text"], font=("Segoe UI", 13))

        # Entries and Comboboxes (focus behaviors will be controlled by customtkinter CTkEntry/CTkComboBox)
        self.style.configure("TEntry",
                              fieldbackground=self.colors["card"],
                              foreground=self.colors["text"],
                              bordercolor=self.colors["border"],
                              lightcolor=self.colors["border"])

        self.style.configure("TCombobox",
                              fieldbackground=self.colors["card"],
                              foreground=self.colors["text"],
                              bordercolor=self.colors["border"],
                              lightcolor=self.colors["border"],
                              arrowcolor=self.colors["text"])

        # Checkbuttons
        self.style.configure("TCheckbutton",
                              background=self.colors["bg"],
                              foreground=self.colors["text"])

        # Treeview Styles with flat headers and custom row padding
        self.style.configure("Treeview",
                              background=self.colors["card"],
                              foreground=self.colors["text"],
                              fieldbackground=self.colors["card"],
                              rowheight=28,
                              borderwidth=0,
                              font=("Segoe UI", 13))

        self.style.configure("Treeview.Heading",
                              background=self.colors["border"],
                              foreground=self.colors["text_muted"],
                              bordercolor=self.colors["border"],
                              lightcolor=self.colors["border"],
                              darkcolor=self.colors["border"],
                              borderwidth=1,
                              font=("Segoe UI", 13, "bold"))
        
        self.style.map("Treeview",
                       background=[("selected", self.colors["primary"])],
                       foreground=[("selected", "#ffffff")])
                       
        # Buttons
        if theme == "dark":
            self.style.configure("TButton", 
                                 background=self.colors["primary"], 
                                 foreground="#ffffff", 
                                 bordercolor=self.colors["border"])
            self.style.map("TButton",
                           background=[("active", self.colors["accent"])])
        else:
            self.style.configure("TButton", 
                                 background="#e5e7eb", 
                                 foreground="#1f2937", 
                                 bordercolor="#d1d5db")
            self.style.map("TButton",
                           background=[("active", "#d1d5db")])

        # Update specific Treeview alternating row colors if they exist
        theme = self.config_data.get("theme", "light")
        even_bg = "#121212" if theme == "dark" else "#f9fafb"
        odd_bg  = "#000000" if theme == "dark" else "#ffffff"
        text_fg = "#f9fafb" if theme == "dark" else "#1f2937"

        if hasattr(self, 'tree') and self.tree:
            self.tree.tag_configure('evenrow', background=even_bg, foreground=text_fg)
            self.tree.tag_configure('oddrow', background=odd_bg, foreground=text_fg)
            
        if hasattr(self, 'ledger_tree') and self.ledger_tree:
            self.ledger_tree.tag_configure('evenrow', background=even_bg, foreground=text_fg)
            self.ledger_tree.tag_configure('oddrow', background=odd_bg, foreground=text_fg)
            
        if hasattr(self, 'reports_tab') and hasattr(self.reports_tab, 'tree'):
            self.reports_tab.tree.tag_configure('evenrow', background=even_bg, foreground=text_fg)
            self.reports_tab.tree.tag_configure('oddrow', background=odd_bg, foreground=text_fg)

        if hasattr(self, 'setup_lists_tab') and hasattr(self.setup_lists_tab, 'list_trees'):
            for t in self.setup_lists_tab.list_trees:
                t.tag_configure('evenrow', background=even_bg, foreground=text_fg)
                t.tag_configure('oddrow', background=odd_bg, foreground=text_fg)


        # Traverse and update colors for standard widgets recursively
        self.update_widget_colors(self)
        
        # Trigger dashboard redraw if initialized to update matplotlib colors
        if hasattr(self, "ent_dash_from_year") and self.ent_dash_from_year:
            try:
                self.render_dashboard()
            except Exception:
                pass

    def create_menu_bar(self):
        """
        Creates and mounts the top menu bar containing File, Dashboard, Entries, Transactions Report, Help, and Theme Toggle.
        """
        self.menu_bar = tk.Menu(self)
        
        # File Menu
        self.file_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.file_menu.add_command(label="New Workbook...", command=self.create_new_workbook)
        self.file_menu.add_command(label="Open Workbook...", command=self.open_workbook)
        self.file_menu.add_command(label="Close Workbook", command=self.close_workbook)
        self.file_menu.add_separator()
        self.file_menu.add_command(label="Preferences...", command=self.show_preferences)
        self.file_menu.add_separator()
        self.file_menu.add_command(label="Exit", command=self.destroy)
        self.menu_bar.add_cascade(label="File", menu=self.file_menu)
        
        # Dashboard Menu
        self.menu_bar.add_command(label="Dashboard", command=self.show_dashboard)
        
        # Entries Menu
        self.entries_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.entries_menu.add_command(label="Transaction Entries", command=self.show_transaction_entries)
        self.entries_menu.add_command(label="Ledger Budget", command=self.show_ledger_budget)
        self.entries_menu.add_command(label="Setup Lists...", command=self.show_setup_lists)
        self.menu_bar.add_cascade(label="Entries", menu=self.entries_menu)
        
        # Transactions Report Menu
        self.menu_bar.add_command(label="Transactions Report", command=self.show_transactions_report)
        
        # Help Menu
        self.help_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.help_menu.add_command(label="About", command=self.show_about_dialog)
        self.menu_bar.add_cascade(label="Help", menu=self.help_menu)
        
        # Switch Theme Toggle (Icon)
        theme = self.theme_var.get()
        self.theme_icon = "🌙" if theme == "light" else "☀️"
        self.menu_bar.add_command(label=self.theme_icon, command=self.toggle_theme_from_menu)
        
        # Mount the Menu Bar to the window
        self.configure(menu=self.menu_bar)

    def toggle_theme_from_menu(self):
        """Toggles theme from the main menu icon and updates the icon."""
        current = self.theme_var.get()
        new_theme = "dark" if current == "light" else "light"
        old_icon = "🌙" if current == "light" else "☀️"
        new_icon = "🌙" if new_theme == "light" else "☀️"
        
        self.switch_theme(new_theme)
        
        if hasattr(self, 'menu_bar'):
            try:
                self.menu_bar.entryconfigure(old_icon, label=new_icon)
            except Exception:
                pass

    def show_preferences(self):
        """Displays the Preferences screen."""
        self._previous_screen = None
        if hasattr(self, 'transaction_tab') and self.transaction_tab.grid_info(): self._previous_screen = self.show_transaction_entries
        elif hasattr(self, 'ledger_tab') and self.ledger_tab.grid_info(): self._previous_screen = self.show_ledger_budget
        elif hasattr(self, 'dashboard_tab') and self.dashboard_tab.grid_info(): self._previous_screen = self.show_dashboard
        elif hasattr(self, 'reports_tab') and self.reports_tab.grid_info(): self._previous_screen = self.show_transactions_report
        elif hasattr(self, 'setup_lists_tab') and self.setup_lists_tab.grid_info(): self._previous_screen = self.show_setup_lists

        self.hide_all_frames()
        self.set_active_nav("preferences")
        if hasattr(self, 'screen_title_label'):
            self.screen_title_label.configure(text="Preferences")
        if hasattr(self, 'preferences_tab'):
            self.preferences_tab.grid(row=1, column=0, sticky="nsew")

    def create_new_workbook(self):
        """
        Prompts the user for a 4-digit year, creates a customized Excel workbook
        with 'TransactionEntries' and 'LedgerBudget' worksheets, writes headers,
        configures column widths, and saves it.
        """
        if openpyxl is None:
            messagebox.showerror(
                "Error", 
                "The 'openpyxl' library is required to create Excel workbooks.\n"
                "Please run: pip install openpyxl"
            )
            return

        # Prompt for year
        dialog = CustomInputDialog(self, "New Workbook", "Enter Year (4-digit integer):", theme_colors=self.theme_colors)
        
        # If user cancels or closes dialog
        if dialog.result is None or str(dialog.result).strip() == "":
            return
            
        try:
            year = int(dialog.result)
            if not (1000 <= year <= 9999):
                messagebox.showerror("Invalid Input", "Please enter a valid 4-digit year (e.g., 2026).", parent=self)
                return
        except ValueError:
            messagebox.showerror("Invalid Input", "Please enter a valid integer for the year.", parent=self)
            return

        # Determine target directory
        dirs = self.config_data.get("directories", {})
        target_dir = dirs.get("working_directory", "").strip()
        if not target_dir or not os.path.isdir(target_dir):
            target_dir = os.getcwd()
            
        filename = os.path.join(target_dir, f"{year}-FinancialRecords.xlsx")
        
        # Check if file already exists to avoid accidental overwriting
        if os.path.exists(filename):
            messagebox.showerror(
                "File Exists", 
                f"A workbook named '{os.path.basename(filename)}' already exists in the working directory.\nPlease choose a different year or move the existing file.",
                parent=self
            )
            return

        try:
            # Create a new workbook
            wb = openpyxl.Workbook()
            
            # Create the worksheets we need
            ws_trans = wb.create_sheet(title="TransactionEntries")
            ws_ledger = wb.create_sheet(title="LedgerBudget")
            
            # Delete default sheet (usually named "Sheet")
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
                
            # Define headers
            trans_headers = [
                "Date", "Voucher Number", "Type", "Source Type", 
                "Source Ref", "Description", "Ledger", "Financial Report Category", "Amount"
            ]
            ledger_headers = ["Ledger", "Month", "Year", "Amount"]
            
            # Write headers
            ws_trans.append(trans_headers)
            ws_ledger.append(ledger_headers)
            
            # Formatting (make headers bold)
            bold_font = openpyxl.styles.Font(bold=True)
            
            for col_idx in range(1, len(trans_headers) + 1):
                cell = ws_trans.cell(row=1, column=col_idx)
                cell.font = bold_font
            
            for col_idx in range(1, len(ledger_headers) + 1):
                cell = ws_ledger.cell(row=1, column=col_idx)
                cell.font = bold_font

            # Auto-adjust column widths with a sensible minimum
            trans_widths = {
                1: 12,  # Date
                2: 18,  # Voucher Number
                3: 10,  # Type
                4: 15,  # Source Type
                5: 15,  # Source Ref
                6: 30,  # Description
                7: 15,  # Ledger
                8: 12   # Amount
            }
            for col_idx, width in trans_widths.items():
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                ws_trans.column_dimensions[col_letter].width = width

            ledger_widths = {
                1: 15,  # Ledger
                2: 10,  # Month
                3: 10,  # Year
                4: 12   # Amount
            }
            for col_idx, width in ledger_widths.items():
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                ws_ledger.column_dimensions[col_letter].width = width

            # Save the workbook
            wb.save(filename)
            
            # Auto-open the newly created workbook
            self.current_workbook_path = os.path.abspath(filename)
            self.is_workbook_valid = True
            self.update_app_title()
            
            self.status_label.configure(
                text=f"Active Workbook: {filename}",
                fg=self.colors["accent"]
            )
            
            self.set_forms_state("normal")
            self.load_workbook_data()
            self.show_dashboard()
            
            messagebox.showinfo(
                "Success", 
                f"Successfully created workbook:\n{filename}",
                parent=self
            )
            
        except Exception as e:
            messagebox.showerror(
                "Error", 
                f"An error occurred while creating the workbook:\n{e}",
                parent=self
            )

    def init_config(self):
        """
        Ensures the config.json file exists. If not, creates it with default values.
        Then loads the configuration.
        """
        self.config_file = "config.json"
        self.default_config = {
            "types": ["Payment", "Receipt", "Journal"],
            "source_types": ["Bank", "Cash", "Credit Card"],
            "ledgers": ["Sales", "Purchases", "Expenses", "Income"],
            "fin_report_categories": ["Operating", "Investing", "Financing"],
            "ledger_category_mapping": {
                "Sales": "Operating",
                "Purchases": "Operating",
                "Expenses": "Operating",
                "Income": "Operating"
            },
            "theme": "light"
        }
        
        if not os.path.exists(self.config_file):
            self.save_config(self.default_config)
            self.config_data = self.default_config
        else:
            self.load_config()

    def update_app_title(self):
        """Updates the application window title based on open workbook."""
        title_base = "Financial Records Manager"
        if self.current_workbook_path:
            filename = os.path.basename(self.current_workbook_path)
            invalid_suffix = " [Invalid]" if not getattr(self, "is_workbook_valid", True) else ""
            self.title(f"{title_base} - {filename}{invalid_suffix}")
        else:
            self.title(title_base)

    def load_config(self):
        """Loads configuration from config.json."""
        try:
            with open(self.config_file, "r") as f:
                self.config_data = json.load(f)
            # Ensure all required keys exist
            for key, val in self.default_config.items():
                if key not in self.config_data:
                    self.config_data[key] = val
        except Exception as e:
            messagebox.showerror("Error", f"Could not load configuration file:\n{e}", parent=self)
            self.config_data = self.default_config.copy()

    def save_config(self, data=None):
        """Saves configuration to config.json."""
        if data is None:
            data = self.config_data
        try:
            with open(self.config_file, "w") as f:
                json.dump(data, f, indent=4)
        except Exception as e:
            messagebox.showerror("Error", f"Could not save configuration file:\n{e}", parent=self)

    def update_config_list(self, key, new_list):
        """Updates a configuration list immediately and writes to JSON."""
        if key in self.config_data:
            self.config_data[key] = new_list
            self.save_config()
            self.refresh_comboboxes()

    def show_setup_lists(self):
        """Displays the Setup Lists screen."""
        self.hide_all_frames()
        self.set_active_nav("lists")
        if hasattr(self, 'screen_title_label'):
            self.screen_title_label.configure(text="Setup Lists")
        self.setup_lists_tab.grid(row=1, column=0, sticky="nsew")

    def open_workbook(self):
        """
        Prompts the user to select an Excel workbook.
        Performs validation checks and loads data if valid.
        """
        if openpyxl is None:
            messagebox.showerror(
                "Error",
                "The 'openpyxl' library is required to read Excel files.\n"
                "Please run: pip install openpyxl"
            )
            return

        working_dir = self.config_data.get("directories", {}).get("working_directory", "")
        if working_dir and not os.path.isdir(working_dir):
            working_dir = ""

        filepath = filedialog.askopenfilename(
            initialdir=working_dir if working_dir else None,
            title="Open Workbook",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
            parent=self
        )
        if not filepath:
            return

        filename = os.path.basename(filepath)
        
        # Check name pattern: YYYY-FinancialRecords.xlsx
        import re
        is_pattern_match = re.match(r"^\d{4}-FinancialRecords\.xlsx$", filename)
        if not is_pattern_match:
            messagebox.showwarning("Warning", "File naming convention not followed.", parent=self)

        # Clear any old data
        for item in self.tree.get_children():
            self.tree.delete(item)
        if hasattr(self, 'ledger_tree'):
            for item in self.ledger_tree.get_children():
                self.ledger_tree.delete(item)

        # Try to open and validate workbook structure
        try:
            wb = openpyxl.load_workbook(filepath)
            
            # Check sheet names
            required_sheets = ["TransactionEntries", "LedgerBudget"]
            structure_valid = True
            
            for sheet in required_sheets:
                if sheet not in wb.sheetnames:
                    structure_valid = False
                    break
            
            if structure_valid:
                # Check column headers for TransactionEntries
                ws_trans = wb["TransactionEntries"]
                legacy_expected_trans_headers = [
                    "Date", "Voucher Number", "Type", "Source Type", 
                    "Source Ref", "Description", "Ledger", "Amount"
                ]
                expected_trans_headers = [
                    "Date", "Voucher Number", "Type", "Source Type", 
                    "Source Ref", "Description", "Ledger", "Financial Report Category", "Amount"
                ]
                trans_headers = [cell.value for cell in ws_trans[1]]
                
                # Check for legacy and migrate
                if len(trans_headers) == len(legacy_expected_trans_headers) and \
                   trans_headers[:len(legacy_expected_trans_headers)] == legacy_expected_trans_headers:
                    ws_trans.insert_cols(8) # Insert before Amount (index 8, 1-based)
                    ws_trans.cell(row=1, column=8).value = "Financial Report Category"
                    ws_trans.cell(row=1, column=8).font = openpyxl.styles.Font(bold=True)
                    wb.save(filepath)
                    trans_headers = [cell.value for cell in ws_trans[1]]

                if len(trans_headers) < len(expected_trans_headers) or \
                   trans_headers[:len(expected_trans_headers)] != expected_trans_headers:
                    structure_valid = False

            if structure_valid:
                # Check column headers for LedgerBudget
                ws_ledger = wb["LedgerBudget"]
                expected_ledger_headers = ["Ledger", "Month", "Year", "Amount"]
                ledger_headers = [cell.value for cell in ws_ledger[1]]
                if len(ledger_headers) < len(expected_ledger_headers) or \
                   ledger_headers[:len(expected_ledger_headers)] != expected_ledger_headers:
                    structure_valid = False

            if not structure_valid:
                # Store filepath to track active selection but mark as invalid
                self.current_workbook_path = filepath
                self.is_workbook_valid = False
                self.update_app_title()
                
                # Show red status label
                self.status_label.configure(
                    text="Invalid sheet structure. Cannot proceed.",
                    fg="#dc2626"  # Red
                )
                
                # Disable form inputs
                self.set_forms_state("disabled")
                
                messagebox.showerror(
                    "Error", 
                    "Invalid workbook structure. The required sheets or headers do not match. Forms have been disabled.",
                    parent=self
                )
                return

            # If validation is successful, store path and load
            self.current_workbook_path = filepath
            # Update title to include opened filename
            self.is_workbook_valid = True
            self.update_app_title()
            
            self.status_label.configure(
                text=f"Active Workbook: {filename}",
                fg=self.colors["accent"]
            )
            
            # Enable forms
            self.set_forms_state("normal")
            
            self.load_workbook_data()
            self.show_dashboard()
            
        except Exception as e:
            self.current_workbook_path = None
            self.status_label.configure(
                text="No workbook open. Create or open one to proceed.",
                fg=self.colors["text_muted"]
            )
            self.set_forms_state("disabled")
            
            messagebox.showerror(
                "Error",
                f"An error occurred while reading the workbook:\n{e}",
                parent=self
            )

    def load_workbook_data(self):
        """
        Reads rows from 'TransactionEntries' and 'LedgerBudget' sheets
        and renders them into their respective Treeviews.
        """
        # Clear transaction treeview items
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Clear ledger treeview items
        if hasattr(self, 'ledger_tree'):
            for item in self.ledger_tree.get_children():
                self.ledger_tree.delete(item)

        if not hasattr(self, 'current_workbook_path') or not self.current_workbook_path:
            return

        try:
            wb = openpyxl.load_workbook(self.current_workbook_path, data_only=True)
            
            # Load TransactionEntries
            ws_trans = wb["TransactionEntries"]
            trans_records = []
            for row_idx in range(2, ws_trans.max_row + 1):
                row_vals = [ws_trans.cell(row=row_idx, column=c).value for c in range(1, 10)]
                
                if all(val is None for val in row_vals):
                    continue
                
                if isinstance(row_vals[0], datetime):
                    row_vals[0] = row_vals[0].strftime("%d-%m-%Y")
                elif row_vals[0] is not None:
                    val_str = str(row_vals[0]).strip()
                    try:
                        # Detect and format legacy YYYY-MM-DD strings
                        dt = datetime.strptime(val_str, "%Y-%m-%d")
                        row_vals[0] = dt.strftime("%d-%m-%Y")
                    except ValueError:
                        row_vals[0] = val_str
                else:
                    row_vals[0] = ""
                    
                if row_vals[8] is not None:
                    try:
                        row_vals[8] = f"{float(row_vals[8]):,.2f}"
                    except ValueError:
                        pass
                
                date_val = str(row_vals[0] or "").strip()
                try:
                    sort_date = datetime.strptime(date_val, "%d-%m-%Y").strftime("%Y-%m-%d")
                except ValueError:
                    sort_date = date_val

                try:
                    sort_voucher = float(row_vals[1])
                except (ValueError, TypeError):
                    sort_voucher = float('inf')

                record = (
                    row_idx,
                    row_vals[0],
                    row_vals[1],
                    row_vals[2],
                    row_vals[3],
                    row_vals[4],
                    row_vals[5],
                    row_vals[6],
                    row_vals[7],
                    row_vals[8],
                    sort_voucher,           # sort_voucher
                    sort_date               # sort_date
                )
                trans_records.append(record)
                
            # Sort by Voucher Number descending, then Date descending
            trans_records.sort(key=lambda r: (r[10], r[11]), reverse=True)
            
            for index, r in enumerate(trans_records):
                row_tag = 'evenrow' if index % 2 == 0 else 'oddrow'
                self.tree.insert(
                    "", 
                    tk.END, 
                    values=(r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7], r[8], r[9]),
                    tags=(row_tag,)
                )

            if hasattr(self, 'lbl_trans_status'):
                self.lbl_trans_status.configure(text=f"Total Transactions: {len(trans_records)}")

            # Load LedgerBudget
            if hasattr(self, 'ledger_tree'):
                selected_ledger_filter = self.cb_l_ledger.get().strip() if hasattr(self, 'cb_l_ledger') else "<None>"
                ws_ledger = wb["LedgerBudget"]
                ledger_records = []
                for row_idx in range(2, ws_ledger.max_row + 1):
                    if selected_ledger_filter == "<None>":
                        break
                        
                    row_vals = [ws_ledger.cell(row=row_idx, column=c).value for c in range(1, 5)]
                    
                    if all(val is None for val in row_vals):
                        continue
                        
                    if selected_ledger_filter != "<All Ledgers>":
                        if str(row_vals[0]).strip() != selected_ledger_filter:
                            continue
                    
                    m_num = 0
                    if row_vals[1] is not None:
                        import calendar
                        try:
                            # If it's already an integer:
                            m_idx = int(row_vals[1])
                            if 1 <= m_idx <= 12:
                                m_num = m_idx
                                row_vals[1] = calendar.month_name[m_idx]
                        except ValueError:
                            # If it's already a string, map it back to int for sorting
                            m_str = str(row_vals[1]).strip().lower()
                            for i, m_name in enumerate(calendar.month_name):
                                if m_name and m_name.lower() == m_str:
                                    m_num = i
                                    break

                    if row_vals[3] is not None:
                        try:
                            row_vals[3] = f"{float(row_vals[3]):,.2f}"
                        except ValueError:
                            pass
                    
                    record = (
                        row_idx,
                        row_vals[0],
                        row_vals[1],
                        row_vals[2],
                        row_vals[3],
                        str(row_vals[0] or "").lower(), # sort_ledger
                        m_num,                          # sort_month
                        int(row_vals[2]) if row_vals[2] and str(row_vals[2]).strip().isdigit() else 0 # sort_year
                    )
                    ledger_records.append(record)
                    
                # Sort by Ledger, then Year, then Month Number
                ledger_records.sort(key=lambda r: (r[5], r[7], r[6]))
                
                for index, r in enumerate(ledger_records):
                    row_tag = 'evenrow' if index % 2 == 0 else 'oddrow'
                    self.ledger_tree.insert(
                        "",
                        tk.END,
                        values=(r[0], r[1], r[2], r[3], r[4]),
                        tags=(row_tag,)
                    )
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load workbook data:\n{e}", parent=self)

    def check_workbook_open(self):
        """Checks if a workbook path is set, otherwise issues a warning."""
        if not hasattr(self, 'current_workbook_path') or not self.current_workbook_path:
            messagebox.showwarning("No Workbook Open", "Please open or create a workbook first.", parent=self)
            return False
        return True

    def validate_inputs(self):
        """
        Validates form inputs. Returns a dictionary of validated values if success,
        otherwise shows error boxes and returns None.
        """
        date_str = self.ent_date.get().strip()
        voucher_str = self.ent_voucher.get().strip()
        type_str = self.cb_type.get().strip()
        source_type_str = self.cb_source_type.get().strip()
        source_ref_str = self.ent_source_ref.get().strip()
        desc_str = self.ent_desc.get().strip().title()
        ledger_str = self.cb_ledger.get().strip()
        amount_str = self.ent_amount.get().strip()
        
        if not date_str or not voucher_str or not type_str or not source_type_str or not desc_str or not ledger_str or not amount_str:
            messagebox.showerror("Validation Error", "All fields except Source Ref are required.", parent=self)
            return None
            
        try:
            parsed_date = datetime.strptime(date_str, "%d-%m-%Y")
        except ValueError:
            messagebox.showerror("Validation Error", "Date must be in DD-MM-YYYY format.", parent=self)
            return None
            
        try:
            amount_val = float(amount_str.replace(",", ""))
            if amount_val < 0:
                messagebox.showerror("Validation Error", "Amount cannot be negative.", parent=self)
                return None
        except ValueError:
            messagebox.showerror("Validation Error", "Amount must be a valid numeric value.", parent=self)
            return None
            
        fin_category = self.config_data.get("ledger_category_mapping", {}).get(ledger_str, "")

        return {
            "date": parsed_date.strftime("%Y-%m-%d"),
            "voucher": voucher_str,
            "type": type_str,
            "source_type": source_type_str,
            "source_ref": source_ref_str,
            "desc": desc_str,
            "ledger": ledger_str,
            "fin_category": fin_category,
            "amount": amount_val
        }

    def add_transaction(self):
        """Validates input data and appends a new transaction row to the Excel workbook."""
        if not self.check_workbook_open():
            return
            
        vals = self.validate_inputs()
        if not vals:
            return
            
        try:
            wb = openpyxl.load_workbook(self.current_workbook_path)
            ws = wb["TransactionEntries"]
            
            new_row = [
                vals["date"],
                vals["voucher"],
                vals["type"],
                vals["source_type"],
                vals["source_ref"],
                vals["desc"],
                vals["ledger"],
                vals["fin_category"],
                vals["amount"]
            ]
            ws.append(new_row)
            
            wb.save(self.current_workbook_path)
            
            self.load_workbook_data()
            self.clear_form()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to add transaction:\n{e}", parent=self)

    def retrieve_transaction(self):
        """Loads data from the selected Treeview item back into the form fields."""
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showwarning("No Selection", "Please select a record from the grid to retrieve.", parent=self)
            return
            
        vals = self.tree.item(selected_item, "values")
        if not vals:
            return
            
        self.clear_form()
        self.ent_date.insert(0, vals[1])
        self.ent_voucher.insert(0, vals[2])
        self.cb_type.set(vals[3])
        self.cb_source_type.set(vals[4])
        self.ent_source_ref.insert(0, vals[5] if vals[5] is not None and vals[5] != "None" else "")
        self.ent_desc.insert(0, vals[6])
        self.cb_ledger.set(vals[7])
        try:
            amt_float = float(str(vals[9]).replace(",", ""))
            self.ent_amount.insert(0, f"{amt_float:,.2f}")
        except (ValueError, TypeError):
            self.ent_amount.insert(0, vals[9])

    def update_transaction(self):
        """Updates the selected transaction row in the Excel workbook with current form data."""
        if not self.check_workbook_open():
            return
            
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showwarning("No Selection", "Please select a record from the grid to update.", parent=self)
            return
            
        vals = self.validate_inputs()
        if not vals:
            return
            
        tree_vals = self.tree.item(selected_item, "values")
        excel_row_idx = int(tree_vals[0])
        
        try:
            wb = openpyxl.load_workbook(self.current_workbook_path)
            ws = wb["TransactionEntries"]
            
            updated_vals = [
                vals["date"],
                vals["voucher"],
                vals["type"],
                vals["source_type"],
                vals["source_ref"],
                vals["desc"],
                vals["ledger"],
                vals["fin_category"],
                vals["amount"]
            ]
            
            for col_idx, val in enumerate(updated_vals, 1):
                ws.cell(row=excel_row_idx, column=col_idx, value=val)
                
            wb.save(self.current_workbook_path)
            
            self.load_workbook_data()
            self.clear_form()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update transaction:\n{e}", parent=self)

    def delete_transaction(self):
        """Deletes the selected transaction row from the Excel workbook."""
        if not self.check_workbook_open():
            return
            
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showwarning("No Selection", "Please select a record from the grid to delete.", parent=self)
            return
            
        tree_vals = self.tree.item(selected_item, "values")
        excel_row_idx = int(tree_vals[0])
        voucher_num = tree_vals[2]
        
        confirm = messagebox.askyesno(
            "Confirm Delete", 
            f"Are you sure you want to delete transaction with Voucher Number '{voucher_num}'?", 
            parent=self
        )
        if not confirm:
            return
            
        try:
            wb = openpyxl.load_workbook(self.current_workbook_path)
            ws = wb["TransactionEntries"]
            
            ws.delete_rows(excel_row_idx)
            
            wb.save(self.current_workbook_path)
            
            self.load_workbook_data()
            self.clear_form()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to delete transaction:\n{e}", parent=self)

    def clear_form(self):
        """Clears all input fields in the Transaction Entry form."""
        self.ent_date.delete(0, tk.END)
        self.ent_voucher.delete(0, tk.END)
        self.cb_type.set("")
        self.cb_source_type.set("")
        self.ent_source_ref.delete(0, tk.END)
        self.ent_desc.delete(0, tk.END)
        self.cb_ledger.set("")
        self.ent_amount.delete(0, tk.END)

    def on_tree_select(self, event=None):
        """Enables the Print Voucher button if a row is selected and a workbook is open, otherwise disables it."""
        if hasattr(self, 'btn_print_voucher') and self.btn_print_voucher:
            if self.current_workbook_path and self.tree.selection():
                self.btn_print_voucher.configure(state="normal")
            else:
                self.btn_print_voucher.configure(state="disabled")

    def print_voucher(self):
        """Generates a clean, professional transaction voucher PDF using fpdf2 and opens it."""
        if not self.check_workbook_open():
            return

        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showwarning("No Selection", "Please select a transaction record to print.", parent=self)
            return

        selected_vals = self.tree.item(selected_item, "values")
        if not selected_vals:
            return

        voucher_no = selected_vals[2]
        
        # Aggregate all matching items
        matching_items = []
        if voucher_no and str(voucher_no).strip():
            for child_id in self.tree.get_children():
                child_vals = self.tree.item(child_id, "values")
                if child_vals[2] == voucher_no:
                    matching_items.append(child_vals)
        else:
            matching_items.append(selected_vals)
            
        # Extract metadata from the first matched item
        first_vals = matching_items[0]
        date_val = first_vals[1]
        type_val = first_vals[3]
        source_val = first_vals[4]
        source_ref = first_vals[5]

        if fpdf is None:
            messagebox.showerror(
                "Error",
                "The 'fpdf2' library is required to print vouchers.\n"
                "Please run: pip install fpdf2",
                parent=self
            )
            return

        # Generate output folder paths
        dirs = self.config_data.get("directories", {})
        vouchers_dir = dirs.get("vouchers_directory", "").strip()
        if not vouchers_dir or not os.path.isdir(vouchers_dir):
            wb_dir = os.path.dirname(self.current_workbook_path)
            vouchers_dir = os.path.join(wb_dir, "vouchers")
        
        os.makedirs(vouchers_dir, exist_ok=True)
        pdf_filename = f"Voucher_{voucher_no}.pdf"
        pdf_path = os.path.join(vouchers_dir, pdf_filename)

        try:
            company_info = self.config_data.get("company_info", {})
            pdf = VoucherPDF(company_info=company_info, orientation="P", unit="mm", format="A4")
            pdf.set_margins(20, 20, 20)
            pdf.add_page()

            # Metadata Section Grid
            # We want to use comfortable line heights
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_text_color(75, 85, 99)
            pdf.cell(35, 8, "Voucher Number:")
            pdf.set_font("Helvetica", "", 10)
            pdf.set_text_color(31, 41, 55)
            pdf.cell(55, 8, str(voucher_no))

            pdf.set_font("Helvetica", "B", 10)
            pdf.set_text_color(75, 85, 99)
            pdf.cell(35, 8, "Transaction Date:")
            pdf.set_font("Helvetica", "", 10)
            pdf.set_text_color(31, 41, 55)
            pdf.cell(45, 8, str(date_val), new_x="LMARGIN", new_y="NEXT")

            # Row 2
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_text_color(75, 85, 99)
            pdf.cell(35, 8, "Voucher Type:")
            pdf.set_font("Helvetica", "", 10)
            pdf.set_text_color(31, 41, 55)
            pdf.cell(55, 8, str(type_val))

            pdf.set_font("Helvetica", "B", 10)
            pdf.set_text_color(75, 85, 99)
            pdf.cell(35, 8, "Source Type:")
            pdf.set_font("Helvetica", "", 10)
            pdf.set_text_color(31, 41, 55)
            pdf.cell(45, 8, str(source_val), new_x="LMARGIN", new_y="NEXT")

            # Row 3
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_text_color(75, 85, 99)
            pdf.cell(35, 8, "Source Ref:")
            pdf.set_font("Helvetica", "", 10)
            pdf.set_text_color(31, 41, 55)
            pdf.cell(55, 8, str(source_ref) if source_ref and source_ref != "None" else "N/A", new_x="LMARGIN", new_y="NEXT")

            pdf.ln(10)

            # Details Table Header
            pdf.set_fill_color(30, 58, 138)  # Primary dark blue
            pdf.set_text_color(255, 255, 255) # White
            pdf.set_draw_color(30, 58, 138)
            pdf.set_line_width(0.3)
            pdf.set_font("Helvetica", "B", 10)
            pdf.cell(10, 10, " S/N", border=1, fill=True, align="L")
            pdf.cell(75, 10, " Description", border=1, fill=True, align="L")
            pdf.cell(40, 10, " Ledger Category", border=1, fill=True, align="L")
            
            font_bold_path = "C:\\Windows\\Fonts\\arialbd.ttf"
            amount_header = "Amount (NGN) "
            if os.path.exists(font_bold_path):
                try:
                    import warnings
                    with warnings.catch_warnings():
                        warnings.simplefilter("ignore")
                        pdf.add_font("Arial", "B", font_bold_path)
                    pdf.set_font("Arial", "B", 10)
                    amount_header = "Amount (\u20a6) "
                except Exception:
                    pass
            pdf.cell(45, 10, amount_header, border=1, fill=True, align="R", new_x="LMARGIN", new_y="NEXT")

            # Details Table Value Rows
            pdf.set_text_color(31, 41, 55)
            pdf.set_font("Helvetica", "", 10)
            
            total_amount = 0.0
            sn = 1
            
            for item_vals in matching_items:
                i_ledger = item_vals[7]
                i_desc = item_vals[6]
                i_amt_raw = item_vals[9]
                
                try:
                    i_amt = float(str(i_amt_raw).replace(",", ""))
                except (ValueError, TypeError):
                    i_amt = 0.0
                total_amount += i_amt
                
                x = pdf.get_x()
                y = pdf.get_y()
                
                # We need to find the max height between ledger and description
                pdf.set_xy(x, y)
                pdf.multi_cell(10, 8, str(sn), border=0, align="C")
                sn_y2 = pdf.get_y()
                
                pdf.set_xy(x + 12, y)
                pdf.multi_cell(71, 8, str(i_desc), border=0, align="L")
                desc_y2 = pdf.get_y()
                
                pdf.set_xy(x + 87, y)
                pdf.multi_cell(36, 8, str(i_ledger), border=0, align="L")
                ledger_y2 = pdf.get_y()
                
                max_y = max(sn_y2, desc_y2, ledger_y2)
                h = max_y - y
                
                # Now draw the actual borders
                pdf.set_xy(x, y)
                pdf.cell(10, h, "", border=1)
                pdf.set_xy(x + 10, y)
                pdf.cell(75, h, "", border=1)
                pdf.set_xy(x + 85, y)
                pdf.cell(40, h, "", border=1)
                
                # Position amount cell
                pdf.set_xy(x + 125, y)
                pdf.cell(45, h, f"{i_amt:,.2f} ", border=1, new_x="LMARGIN", new_y="NEXT", align="R")
                
                # Ensure the next row starts at the correct height
                pdf.set_y(max_y)
                sn += 1

            # Draw Totals Row
            pdf.set_font("Helvetica", "B", 10)
            pdf.cell(125, 10, "TOTAL", border=1, align="C")
            pdf.cell(45, 10, f"{total_amount:,.2f} ", border=1, new_x="LMARGIN", new_y="NEXT", align="R")
            
            pdf.ln(6)

            # Amount in Words shaded box
            amount_words = number_to_words(total_amount)
            words_text = f" Amount in Words: {amount_words}"
            
            pdf.set_fill_color(249, 250, 251) # Very light gray background
            pdf.set_draw_color(229, 231, 235)
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_text_color(75, 85, 99)
            
            # Calculate height for amount in words box
            pdf.multi_cell(0, 8, words_text, border=1, fill=True, align="L")
            
            pdf.ln(25) # Space for signatures

            # Signatures Block
            sig_y = pdf.get_y()
            
            pdf.set_font("Helvetica", "", 10)
            pdf.set_text_color(75, 85, 99)
            
            pdf.set_xy(20, sig_y)
            pdf.cell(50, 5, "_______________________", align="C")
            
            pdf.set_xy(120, sig_y)
            pdf.cell(50, 5, "_______________________", new_x="LMARGIN", new_y="NEXT", align="C")
            
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_text_color(31, 41, 55)
            
            pdf.set_xy(20, sig_y + 6)
            pdf.cell(50, 5, "Prepared By", align="C")
            
            pdf.set_xy(120, sig_y + 6)
            pdf.cell(50, 5, "Authorized Signature", new_x="LMARGIN", new_y="NEXT", align="C")

            pdf.output(pdf_path)

        except PermissionError:
            messagebox.showerror(
                "Permission Error",
                f"Could not save the PDF file because it is currently open in another application.\n"
                f"Please close the file '{pdf_filename}' and try again.",
                parent=self
            )
            return
        except Exception as e:
            messagebox.showerror(
                "Error",
                f"An error occurred while generating the PDF:\n{e}",
                parent=self
            )
            return

        # Open the PDF using system viewer
        try:
            os.startfile(pdf_path)
        except Exception as e:
            messagebox.showerror(
                "Error",
                f"Failed to open the PDF voucher file:\n{e}",
                parent=self
            )

    def show_about_dialog(self):
        """Displays the 'Help > About' dialog with details about the app."""
        about_win = ctk.CTkToplevel(self)
        about_win.title("About Financial Records Manager")
        about_win.geometry("480x320")
        about_win.resizable(False, False)
        
        # Center the window relative to the main app window
        about_win.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() - 480) // 2
        y = self.winfo_y() + (self.winfo_height() - 320) // 2
        about_win.geometry(f"+{x}+{y}")
        
        # Make it modal
        about_win.transient(self)
        about_win.grab_set()

        title_font = ctk.CTkFont(family="Segoe UI", size=20, weight="bold")
        text_font = ctk.CTkFont(family="Segoe UI", size=13)
        
        lbl_title = ctk.CTkLabel(
            about_win, 
            text="Financial Records Manager\nVersion 1.1.0", 
            font=title_font, 
            justify="center"
        )
        lbl_title.pack(pady=(30, 15))
        
        desc = (
            "A premium accounting and budgeting solution featuring a modernized interface, "
            "interactive dashboard analytics, comprehensive Excel-based transaction tracking, "
            "dynamic year projections, and accessible PDF voucher generation.\n\n"
            "Powered by Python, CustomTkinter, Openpyxl, Matplotlib, and fpdf2."
        )
        lbl_desc = ctk.CTkLabel(
            about_win, 
            text=desc, 
            font=text_font, 
            wraplength=420, 
            justify="center"
        )
        lbl_desc.pack(padx=20, pady=(0, 20))
        
        btn_close = ctk.CTkButton(
            about_win, 
            text="Close", 
            command=about_win.destroy, 
            width=120
        )
        btn_close.pack(pady=(10, 20))

    def pick_date(self):
        CalendarDialog(self, self.ent_date)

    def format_amount_field(self, event=None):
        """Formats the Amount entry text with thousand separators on focus out."""
        val_str = self.ent_amount.get().strip()
        if not val_str:
            return
        val_str_clean = val_str.replace(",", "")
        try:
            val_float = float(val_str_clean)
            formatted = f"{val_float:,.2f}"
            self.ent_amount.delete(0, tk.END)
            self.ent_amount.insert(0, formatted)
        except ValueError:
            pass

    def format_ledger_amount_field(self, event=None):
        """Formats the Ledger Budget Amount entry text with thousand separators on focus out."""
        val_str = self.ent_l_amount.get().strip()
        if not val_str:
            return
        val_str_clean = val_str.replace(",", "")
        try:
            val_float = float(val_str_clean)
            formatted = f"{val_float:,.2f}"
            self.ent_l_amount.delete(0, tk.END)
            self.ent_l_amount.insert(0, formatted)
        except ValueError:
            pass

    def show_transactions_report(self):
        """Displays the Transactions Report screen."""
        if not self.check_workbook_open():
            return
        self.hide_all_frames()
        self.set_active_nav("report")
        if hasattr(self, 'screen_title_label'):
            self.screen_title_label.configure(text="Transactions Report")
        self.reports_tab.grid(row=1, column=0, sticky="nsew")
        self.reports_tab.load_data()

    def refresh_comboboxes(self):
        """Refreshes values in all form Comboboxes using self.config_data."""
        if hasattr(self, 'cb_type') and self.cb_type:
            self.cb_type.configure(values=sorted(self.config_data.get("types", [])))
        if hasattr(self, 'cb_source_type') and self.cb_source_type:
            self.cb_source_type.configure(values=sorted(self.config_data.get("source_types", [])))
        if hasattr(self, 'cb_ledger') and self.cb_ledger:
            self.cb_ledger.configure(values=sorted(self.config_data.get("ledgers", [])))
        if hasattr(self, 'cb_l_ledger') and self.cb_l_ledger:
            ledgers = sorted(self.config_data.get("ledgers", []))
            self.cb_l_ledger.configure(values=["<None>", "<All Ledgers>"] + ledgers)
            if not self.cb_l_ledger.get() or self.cb_l_ledger.get() not in (["<None>", "<All Ledgers>"] + ledgers):
                self.cb_l_ledger.set("<None>")

    def validate_ledger_inputs(self, ignore_row_idx=None):
        """
        Validates LedgerBudget form inputs.
        Ensures fields are correct and checks for duplicates.
        Returns a dictionary of validated values if success, otherwise None.
        """
        import os
        import re
        import calendar

        ledger_str = self.cb_l_ledger.get().strip()
        month_str = self.cb_l_month.get().strip()
        year_str = self.ent_l_year.get().strip()
        amount_str = self.ent_l_amount.get().strip()

        if not ledger_str or not month_str or not year_str or not amount_str:
            messagebox.showerror("Validation Error", "All fields are required.", parent=self)
            return None
            
        if ledger_str in ("<None>", "<All Ledgers>"):
            messagebox.showerror("Validation Error", "Please select a specific ledger to add or update a budget.", parent=self)
            return None

        # Convert month name to integer
        month_val = None
        for i, m_name in enumerate(calendar.month_name):
            if m_name and m_name.lower() == month_str.lower():
                month_val = i
                break
        
        if month_val is None:
            messagebox.showerror("Validation Error", "Please select a valid month name.", parent=self)
            return None

        year_val = int(year_str)

        try:
            amount_val = float(amount_str.replace(",", ""))
            if amount_val < 0:
                messagebox.showerror("Validation Error", "Amount cannot be negative.", parent=self)
                return None
        except ValueError:
            messagebox.showerror("Validation Error", "Amount must be a valid numeric value.", parent=self)
            return None

        try:
            wb = openpyxl.load_workbook(self.current_workbook_path)
            ws = wb["LedgerBudget"]
            
            for row_idx in range(2, ws.max_row + 1):
                if ignore_row_idx is not None and row_idx == ignore_row_idx:
                    continue
                
                row_vals = [ws.cell(row=row_idx, column=c).value for c in range(1, 4)]
                if all(v is None for v in row_vals):
                    continue
                
                r_ledger = str(row_vals[0]).strip()
                r_month = str(row_vals[1]).strip()
                r_year = str(row_vals[2]).strip()
                
                if r_ledger == ledger_str and r_year == str(year_val):
                    try:
                        if int(r_month) == month_val:
                            messagebox.showerror(
                                "Duplicate Entry",
                                f"A budget record already exists for:\n"
                                f"Ledger: {ledger_str}, Month: {month_str}, Year: {year_val}",
                                parent=self
                            )
                            return None
                    except ValueError:
                        if r_month.lower() == month_str.lower():
                            messagebox.showerror(
                                "Duplicate Entry",
                                f"A budget record already exists for:\n"
                                f"Ledger: {ledger_str}, Month: {month_str}, Year: {year_val}",
                                parent=self
                            )
                            return None
                    
        except Exception as e:
            messagebox.showerror("Error", f"Failed to check duplicates:\n{e}", parent=self)
            return None

        return {
            "ledger": ledger_str,
            "month": month_str,
            "year": year_val,
            "amount": amount_val
        }

    def add_ledger_budget(self):
        """Validates input data and appends a new budget record to LedgerBudget."""
        if not self.check_workbook_open():
            return

        vals = self.validate_ledger_inputs()
        if not vals:
            return

        try:
            wb = openpyxl.load_workbook(self.current_workbook_path)
            ws = wb["LedgerBudget"]

            new_row = [
                vals["ledger"],
                vals["month"],
                vals["year"],
                vals["amount"]
            ]
            ws.append(new_row)
            wb.save(self.current_workbook_path)

            last_ledger = vals["ledger"]
            self.clear_ledger_form(reload=False)
            self.cb_l_ledger.set(last_ledger)
            self.load_workbook_data()

        except Exception as e:
            messagebox.showerror("Error", f"Failed to add budget record:\n{e}", parent=self)

    def retrieve_ledger_budget(self):
        """Loads data from the selected LedgerBudget Treeview item back into form fields."""
        selected_item = self.ledger_tree.selection()
        if not selected_item:
            messagebox.showwarning("No Selection", "Please select a record from the grid to retrieve.", parent=self)
            return

        vals = self.ledger_tree.item(selected_item, "values")
        if not vals:
            return

        self.clear_ledger_form(reload=False)
        self.cb_l_ledger.set(vals[1])
        self.cb_l_month.set(vals[2])
        self.ent_l_year.insert(0, vals[3])
        try:
            amt_float = float(str(vals[4]).replace(",", ""))
            self.ent_l_amount.insert(0, f"{amt_float:,.2f}")
        except (ValueError, TypeError):
            self.ent_l_amount.insert(0, vals[4])

    def update_ledger_budget(self):
        """Updates the selected budget row in the Excel workbook with current form data."""
        if not self.check_workbook_open():
            return

        selected_item = self.ledger_tree.selection()
        if not selected_item:
            messagebox.showwarning("No Selection", "Please select a record from the grid to update.", parent=self)
            return

        tree_vals = self.ledger_tree.item(selected_item, "values")
        excel_row_idx = int(tree_vals[0])

        vals = self.validate_ledger_inputs(ignore_row_idx=excel_row_idx)
        if not vals:
            return

        try:
            wb = openpyxl.load_workbook(self.current_workbook_path)
            ws = wb["LedgerBudget"]

            updated_vals = [
                vals["ledger"],
                vals["month"],
                vals["year"],
                vals["amount"]
            ]

            for col_idx, val in enumerate(updated_vals, 1):
                ws.cell(row=excel_row_idx, column=col_idx, value=val)

            wb.save(self.current_workbook_path)

            last_ledger = vals["ledger"]
            self.clear_ledger_form(reload=False)
            self.cb_l_ledger.set(last_ledger)
            self.load_workbook_data()

        except Exception as e:
            messagebox.showerror("Error", f"Failed to update budget record:\n{e}", parent=self)

    def delete_ledger_budget(self):
        """Deletes the selected budget row from the Excel workbook."""
        if not self.check_workbook_open():
            return

        selected_item = self.ledger_tree.selection()
        if not selected_item:
            messagebox.showwarning("No Selection", "Please select a record from the grid to delete.", parent=self)
            return

        tree_vals = self.ledger_tree.item(selected_item, "values")
        excel_row_idx = int(tree_vals[0])
        ledger_name = tree_vals[1]
        month_name = tree_vals[2]

        confirm = messagebox.askyesno(
            "Confirm Delete",
            f"Are you sure you want to delete the budget for '{ledger_name}' in '{month_name}'?",
            parent=self
        )
        if not confirm:
            return

        try:
            wb = openpyxl.load_workbook(self.current_workbook_path)
            ws = wb["LedgerBudget"]

            ws.delete_rows(excel_row_idx)
            wb.save(self.current_workbook_path)

            last_ledger = self.cb_l_ledger.get()
            self.clear_ledger_form(reload=False)
            if hasattr(self, 'cb_l_ledger') and last_ledger:
                self.cb_l_ledger.set(last_ledger)
            self.load_workbook_data()
            messagebox.showinfo("Success", "Budget record deleted successfully.", parent=self)

        except Exception as e:
            messagebox.showerror("Error", f"Failed to delete budget record:\n{e}", parent=self)

    def clear_ledger_form(self, reload=True):
        """Clears the ledger budget form fields."""
        if hasattr(self, 'cb_l_ledger'):
            self.cb_l_ledger.set("<None>")
        self.cb_l_month.set("")
        self.ent_l_year.delete(0, tk.END)
        self.ent_l_amount.delete(0, tk.END)
        if reload:
            self.load_workbook_data()

    def update_menu_state(self, state):
        """
        Updates the state of the sidebar buttons
        based on whether a workbook is currently open and valid.
        """
        buttons = [
            'btn_nav_dashboard',
            'btn_nav_trans',
            'btn_nav_ledger',
            'btn_nav_report',
            'btn_nav_lists',
            'btn_nav_close_wb'
        ]
        for name in buttons:
            if hasattr(self, name):
                btn = getattr(self, name)
                if btn:
                    btn.configure(state=state)

    def set_forms_state(self, state):
        """
        Sets the state of entry inputs, comboboxes, and action buttons in both forms.
        state can be 'normal' or 'disabled'.
        """
        self.update_menu_state(state)
        
        widgets = [
            self.ent_date, self.ent_voucher, self.cb_type, self.cb_source_type,
            self.ent_source_ref, self.ent_desc, self.cb_ledger, self.ent_amount,
            self.btn_add, self.btn_update, self.btn_delete, self.btn_retrieve, self.btn_clear,
            
            self.cb_l_ledger, self.cb_l_month, self.ent_l_year, self.ent_l_amount,
            self.btn_l_add, self.btn_l_update, self.btn_l_delete, self.btn_l_retrieve, self.btn_l_clear
        ]
        
        if hasattr(self, "btn_print_voucher") and self.btn_print_voucher:
            widgets.append(self.btn_print_voucher)
        if hasattr(self, "btn_date_picker") and self.btn_date_picker:
            widgets.append(self.btn_date_picker)
        if hasattr(self, "cb_dash_year") and self.cb_dash_year:
            widgets.append(self.cb_dash_year)
        if hasattr(self, "cb_dash_view_mode") and self.cb_dash_view_mode:
            widgets.append(self.cb_dash_view_mode)
        if hasattr(self, "btn_dash_refresh") and self.btn_dash_refresh:
            widgets.append(self.btn_dash_refresh)
        
        for widget in widgets:
            if not widget:
                continue
            try:
                if isinstance(widget, ttk.Combobox):
                    if state == "disabled":
                        widget.configure(state="disabled")
                    else:
                        widget.configure(state="readonly")
                else:
                    widget.configure(state=state)
            except Exception:
                pass

        if state == "normal":
            self.on_tree_select()

    def close_workbook(self):
        """
        Closes the active workbook, clears all data fields and treeviews,
        and sets forms to disabled.
        """
        self.current_workbook_path = None
        self.is_workbook_valid = True
        self.update_app_title()
        
        self.clear_form()
        self.clear_ledger_form()
        
        for item in self.tree.get_children():
            self.tree.delete(item)
        if hasattr(self, 'ledger_tree'):
            for item in self.ledger_tree.get_children():
                self.ledger_tree.delete(item)
                
        self.set_forms_state("disabled")
        
        self.status_label.configure(
            text="No workbook open. Create or open one to proceed.",
            fg=self.colors["text_muted"]
        )
        
        self.hide_all_frames()
        if hasattr(self, 'screen_title_label'):
            self.screen_title_label.configure(text="")

    def set_active_nav(self, active_key):
        """Highlights the active sidebar navigation button and resets others."""
        for key, btn in self.nav_buttons.items():
            if key == active_key:
                btn.configure(fg_color=self.theme_colors["accent"], text_color="#ffffff")
            else:
                btn.configure(fg_color="transparent", text_color=self.theme_colors["text"])

    def create_main_layout(self):
        """
        Creates the left sidebar navigation and the right content area.
        """

        self.columnconfigure(0, weight=0) # Left Sidebar
        self.columnconfigure(1, weight=1) # Right Content Area
        self.rowconfigure(0, weight=1)

        # 1. LEFT SIDEBAR PANEL
        self.sidebar = ctk.CTkFrame(self, width=220, corner_radius=0, fg_color=self.theme_colors["card"])
        self.sidebar.grid(row=0, column=0, sticky="nsew")
        self.sidebar.grid_rowconfigure(13, weight=1) # Spacer row

        # Sidebar Title
        lbl_logo = ctk.CTkLabel(
            self.sidebar,
            text="Financial Manager",
            font=("Segoe UI", 18, "bold"),
            text_color=self.theme_colors["accent"]
        )
        lbl_logo.grid(row=0, column=0, padx=20, pady=(20, 20), sticky="w")

        # Sidebar Navigation Buttons
        self.btn_nav_dashboard = ctk.CTkButton(
            self.sidebar, text="📊 Dashboard", anchor="w",
            fg_color="transparent", text_color=self.theme_colors["text"],
            hover_color=self.theme_colors["bg"], font=("Segoe UI", 13, "bold"),
            command=self.show_dashboard
        )
        self.btn_nav_dashboard.grid(row=1, column=0, padx=10, pady=5, sticky="ew")

        self.btn_nav_trans = ctk.CTkButton(
            self.sidebar, text="📝 Transaction Entries", anchor="w",
            fg_color="transparent", text_color=self.theme_colors["text"],
            hover_color=self.theme_colors["bg"], font=("Segoe UI", 13, "bold"),
            command=self.show_transaction_entries
        )
        self.btn_nav_trans.grid(row=2, column=0, padx=10, pady=5, sticky="ew")

        self.btn_nav_ledger = ctk.CTkButton(
            self.sidebar, text="💰 Ledger Budget", anchor="w",
            fg_color="transparent", text_color=self.theme_colors["text"],
            hover_color=self.theme_colors["bg"], font=("Segoe UI", 13, "bold"),
            command=self.show_ledger_budget
        )
        self.btn_nav_ledger.grid(row=3, column=0, padx=10, pady=5, sticky="ew")

        self.btn_nav_report = ctk.CTkButton(
            self.sidebar, text="📄 Transactions Report", anchor="w",
            fg_color="transparent", text_color=self.theme_colors["text"],
            hover_color=self.theme_colors["bg"], font=("Segoe UI", 13, "bold"),
            command=self.show_transactions_report
        )
        self.btn_nav_report.grid(row=4, column=0, padx=10, pady=5, sticky="ew")

        self.btn_nav_lists = ctk.CTkButton(
            self.sidebar, text="⚙️ Setup Lists", anchor="w",
            fg_color="transparent", text_color=self.theme_colors["text"],
            hover_color=self.theme_colors["bg"], font=("Segoe UI", 13, "bold"),
            command=self.show_setup_lists
        )
        self.btn_nav_lists.grid(row=5, column=0, padx=10, pady=5, sticky="ew")

        self.btn_nav_preferences = ctk.CTkButton(
            self.sidebar, text="🛠️ Preferences", anchor="w",
            fg_color="transparent", text_color=self.theme_colors["text"],
            hover_color=self.theme_colors["bg"], font=("Segoe UI", 13, "bold"),
            command=self.show_preferences
        )
        self.btn_nav_preferences.grid(row=6, column=0, padx=10, pady=5, sticky="ew")

        # Divider
        lbl_div = ctk.CTkLabel(self.sidebar, text="—" * 20, text_color=self.theme_colors["border"])
        lbl_div.grid(row=7, column=0, padx=10, pady=10, sticky="ew")

        # File Actions
        self.btn_nav_new_wb = ctk.CTkButton(
            self.sidebar, text="➕ New Workbook", anchor="w",
            fg_color="transparent", text_color=self.theme_colors["text"],
            hover_color=self.theme_colors["bg"], font=("Segoe UI", 13, "bold"),
            command=self.create_new_workbook
        )
        self.btn_nav_new_wb.grid(row=8, column=0, padx=10, pady=5, sticky="ew")

        self.btn_nav_open_wb = ctk.CTkButton(
            self.sidebar, text="📂 Open Workbook", anchor="w",
            fg_color="transparent", text_color=self.theme_colors["text"],
            hover_color=self.theme_colors["bg"], font=("Segoe UI", 13, "bold"),
            command=self.open_workbook
        )
        self.btn_nav_open_wb.grid(row=9, column=0, padx=10, pady=5, sticky="ew")

        self.btn_nav_close_wb = ctk.CTkButton(
            self.sidebar, text="❌ Close Workbook", anchor="w",
            fg_color="transparent", text_color=self.theme_colors["text"],
            hover_color=self.theme_colors["bg"], font=("Segoe UI", 13, "bold"),
            command=self.close_workbook
        )
        self.btn_nav_close_wb.grid(row=10, column=0, padx=10, pady=5, sticky="ew")

        self.btn_nav_theme = ctk.CTkButton(
            self.sidebar, text="🌗 Toggle Theme", anchor="w",
            fg_color="transparent", text_color=self.theme_colors["text"],
            hover_color=self.theme_colors["bg"], font=("Segoe UI", 13, "bold"),
            command=self.toggle_theme_from_menu
        )
        self.btn_nav_theme.grid(row=11, column=0, padx=10, pady=5, sticky="ew")

        self.btn_nav_about = ctk.CTkButton(
            self.sidebar, text="ℹ️ About", anchor="w",
            fg_color="transparent", text_color=self.theme_colors["text"],
            hover_color=self.theme_colors["bg"], font=("Segoe UI", 13, "bold"),
            command=self.show_about_dialog
        )
        self.btn_nav_about.grid(row=12, column=0, padx=10, pady=5, sticky="ew")

        self.btn_nav_exit = ctk.CTkButton(
            self.sidebar, text="🚪 Exit", anchor="w",
            fg_color="transparent", text_color=self.theme_colors["text"],
            hover_color=self.theme_colors["bg"], font=("Segoe UI", 13, "bold"),
            command=self.destroy
        )
        self.btn_nav_exit.grid(row=14, column=0, padx=10, pady=20, sticky="ew")

        # Keep a tracking dict for active highlighting
        self.nav_buttons = {
            "dashboard": self.btn_nav_dashboard,
            "trans": self.btn_nav_trans,
            "ledger": self.btn_nav_ledger,
            "report": self.btn_nav_report,
            "lists": self.btn_nav_lists,
            "preferences": self.btn_nav_preferences
        }

        # 2. RIGHT CONTENT AREA
        self.content_panel = ctk.CTkFrame(self, corner_radius=0, fg_color=self.theme_colors["bg"])
        self.content_panel.grid(row=0, column=1, sticky="nsew")
        self.content_panel.columnconfigure(0, weight=1)
        self.content_panel.rowconfigure(0, weight=0) # Status Banner
        self.content_panel.rowconfigure(1, weight=1) # Main Container

        # Status Banner at the top of content panel
        self.status_banner = ctk.CTkFrame(self.content_panel, corner_radius=0, fg_color=self.theme_colors["bg"])
        self.status_banner.grid(row=0, column=0, sticky="ew", padx=20, pady=(15, 5))

        self.status_label = tk.Label(
            self.status_banner,
            text="No workbook open. Create or open one to proceed.",
            bg=self.colors["bg"],
            fg=self.colors["text_muted"],
            font=("Segoe UI", 13, "bold"),
            anchor="w"
        )
        self.status_label.pack(fill="x")

        # Main container holds the actual screen frames
        self.main_container = ctk.CTkFrame(self.content_panel, corner_radius=0, fg_color=self.theme_colors["bg"])
        self.main_container.grid(row=1, column=0, sticky="nsew", padx=20, pady=(0, 20))
        self.main_container.columnconfigure(0, weight=1)
        self.main_container.rowconfigure(1, weight=1)

        # Screen Title Label
        self.screen_title_label = tk.Label(
            self.main_container,
            text="",
            bg=self.colors["bg"],
            fg=self.colors["primary"],
            font=("Segoe UI", 20, "bold"),
            anchor="w"
        )
        self.screen_title_label.grid(row=0, column=0, sticky="ew", pady=(0, 15))

        # Create Screen Frames (using our new customtkinter-based frame subclasses)
        self.transaction_tab = ctk.CTkFrame(self.main_container, fg_color="transparent")
        self.ledger_tab = ctk.CTkFrame(self.main_container, fg_color="transparent")
        self.dashboard_tab = ctk.CTkFrame(self.main_container, fg_color="transparent")
        self.reports_tab = TransactionsReportFrame(self.main_container, self)
        self.setup_lists_tab = SetupListsFrame(self.main_container, self)
        self.preferences_tab = PreferencesFrame(self.main_container, self)

        # Populate frames with boilerplate layout cards
        self.populate_transaction_tab()
        self.populate_ledger_tab()
        self.populate_dashboard_tab()

        # Apply theme tags to dynamically created UI objects like the Treeviews
        self.apply_theme()

        # Show empty screen by default since no workbook is open
        self.hide_all_frames()
        self.screen_title_label.configure(text="")

    def hide_all_frames(self):
        """Hides all main screen frames."""
        self.transaction_tab.grid_forget()
        self.ledger_tab.grid_forget()
        self.dashboard_tab.grid_forget()
        self.reports_tab.grid_forget()
        self.setup_lists_tab.grid_forget()
        if hasattr(self, 'preferences_tab'):
            self.preferences_tab.grid_forget()

    def reset_dashboard_date_filters(self):
        """Resets the dashboard From/To fields to the current system month and year."""
        if hasattr(self, 'cb_dash_from_month') and hasattr(self, 'cb_dash_to_month'):
            months_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
            now = datetime.now()
            current_month = months_names[now.month - 1]
            current_year = str(now.year)
            
            self.cb_dash_from_month.set(current_month)
            self.ent_dash_from_year.set(current_year)
            
            self.cb_dash_to_month.set(current_month)
            self.ent_dash_to_year.set(current_year)

    def show_dashboard(self):
        """Displays the Dashboard screen."""
        self.hide_all_frames()
        self.set_active_nav("dashboard")
        if hasattr(self, 'screen_title_label'):
            self.screen_title_label.configure(text="Dashboard")
        self.dashboard_tab.grid(row=1, column=0, sticky="nsew")
        self.reset_dashboard_date_filters()
        try:
            self.refresh_dashboard()
        except Exception:
            pass

    def show_transaction_entries(self):
        """Displays the Transaction Entries screen."""
        self.hide_all_frames()
        self.set_active_nav("trans")
        if hasattr(self, 'screen_title_label'):
            self.screen_title_label.configure(text="Transaction Entries")
        self.transaction_tab.grid(row=1, column=0, sticky="nsew")

    def show_ledger_budget(self):
        """Displays the Ledger Budget screen."""
        self.hide_all_frames()
        self.set_active_nav("ledger")
        if hasattr(self, 'screen_title_label'):
            self.screen_title_label.configure(text="Ledger Budget")
        self.ledger_tab.grid(row=1, column=0, sticky="nsew")

    def populate_transaction_tab(self):
        """
        Sets up the interactive data entry form, button toolbar, and Treeview grid
        on the 'Transaction Entries' tab.
        """
        # Configure responsive layout weights for the tab
        self.transaction_tab.columnconfigure(0, weight=1)
        self.transaction_tab.rowconfigure(2, weight=1)  # Grid row expands

        # Scroll container wraps the top form
        scroll_container = ctk.CTkFrame(self.transaction_tab, fg_color="transparent")
        scroll_container.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        scroll_container.columnconfigure(0, weight=1)

        # Horizontal Scrollbar at top
        h_scroll_form = ttk.Scrollbar(scroll_container, orient="horizontal")
        h_scroll_form.pack(side="top", fill="x")

        form_canvas = tk.Canvas(scroll_container, bg=self.colors.get("bg", "#f8f9fa"), highlightthickness=0)
        form_canvas.pack(side="top", fill="x", expand=True)

        h_scroll_form.configure(command=form_canvas.xview)
        form_canvas.configure(xscrollcommand=h_scroll_form.set)

        # Form Card Container
        card = ctk.CTkFrame(form_canvas, fg_color=self.theme_colors["card"], corner_radius=8, border_width=1, border_color=self.theme_colors["border"])
        form_window_id = form_canvas.create_window((0, 0), window=card, anchor="nw")

        def on_form_canvas_configure(event):
            req_width = card.winfo_reqwidth()
            if event.width > req_width:
                form_canvas.itemconfig(form_window_id, width=event.width)
            else:
                form_canvas.itemconfig(form_window_id, width=req_width)

        def on_card_configure(event):
            form_canvas.configure(scrollregion=form_canvas.bbox("all"))
            form_canvas.configure(height=card.winfo_reqheight())

        form_canvas.bind("<Configure>", on_form_canvas_configure)
        card.bind("<Configure>", on_card_configure)

        # Divide card into two columns and add a dummy column for extra space
        card.columnconfigure(0, weight=0)
        card.columnconfigure(1, weight=0)
        card.columnconfigure(2, weight=1)

        left_half = ctk.CTkFrame(card, fg_color="transparent")
        left_half.grid(row=0, column=0, padx=(15, 15), pady=15, sticky="nsew")
        left_half.columnconfigure(0, weight=1)

        right_half = ctk.CTkFrame(card, fg_color="transparent")
        right_half.grid(row=0, column=1, padx=(15, 15), pady=15, sticky="nsew")
        right_half.columnconfigure(0, weight=1)

        # ---- LEFT HALF FIELDS ----
        left_top_row = ctk.CTkFrame(left_half, fg_color="transparent")
        left_top_row.grid(row=0, column=0, pady=(0, 10), sticky="ew")

        # Configure three columns for Date, Voucher Number, and Source Ref
        left_top_row.columnconfigure(0, weight=0, minsize=160)
        left_top_row.columnconfigure(1, weight=0, minsize=160)
        left_top_row.columnconfigure(2, weight=0, minsize=160)
        # Dummy column to absorb extra space so fields cluster on the left
        left_top_row.columnconfigure(3, weight=1)

        # Column 0: Date
        lbl_date = ttk.Label(left_top_row, text="Date (DD-MM-YYYY):", style="CardLabel.TLabel")
        lbl_date.grid(row=0, column=0, pady=(5, 2), sticky="w")

        date_container = ctk.CTkFrame(left_top_row, fg_color="transparent")
        date_container.grid(row=1, column=0, pady=(0, 5), sticky="w")

        self.ent_date = ctk.CTkEntry(date_container, font=("Segoe UI", 16), width=125)
        self.ent_date.pack(side="left")

        self.btn_date_picker = ctk.CTkButton(date_container, text="📅", width=35, fg_color=self.theme_colors["accent"], command=self.pick_date)
        self.btn_date_picker.pack(side="left", padx=(5, 0))

        # Column 1: Voucher Number
        lbl_voucher = ttk.Label(left_top_row, text="Voucher Number:", style="CardLabel.TLabel")
        lbl_voucher.grid(row=0, column=1, pady=(5, 2), sticky="w", padx=10)
        self.ent_voucher = ctk.CTkEntry(left_top_row, font=("Segoe UI", 16), width=135)
        self.ent_voucher.grid(row=1, column=1, pady=(0, 5), sticky="w", padx=10)

        # Column 2: Source Ref
        lbl_source_ref = ttk.Label(left_top_row, text="Source Ref:", style="CardLabel.TLabel")
        lbl_source_ref.grid(row=0, column=2, pady=(5, 2), sticky="w", padx=10)
        self.ent_source_ref = ctk.CTkEntry(left_top_row, font=("Segoe UI", 16), width=135)
        self.ent_source_ref.grid(row=1, column=2, pady=(0, 5), sticky="w", padx=10)

        # Description
        lbl_desc = ttk.Label(left_half, text="Description:", style="CardLabel.TLabel")
        lbl_desc.grid(row=1, column=0, pady=(5, 2), sticky="w")
        self.ent_desc = MultilineEntry(left_half, self, height=3, width=60, font=("Segoe UI", 15))
        self.ent_desc.grid(row=2, column=0, pady=(0, 5), sticky="w")

        # Amount (moved below description, width=15, aligned left)
        lbl_amount = ttk.Label(left_half, text="Amount:", style="CardLabel.TLabel")
        lbl_amount.grid(row=3, column=0, pady=(5, 2), sticky="w")
        self.ent_amount = ctk.CTkEntry(left_half, font=("Segoe UI", 16), width=135)
        self.ent_amount.grid(row=4, column=0, pady=(0, 5), sticky="w")
        self.ent_amount.bind("<FocusOut>", self.format_amount_field)

        # ---- RIGHT HALF FIELDS ----
        # Type
        lbl_type = ttk.Label(right_half, text="Type:", style="CardLabel.TLabel")
        lbl_type.grid(row=0, column=0, pady=(5, 2), sticky="w")
        self.cb_type = ctk.CTkComboBox(right_half, font=("Segoe UI", 16), width=400)
        self.cb_type.grid(row=1, column=0, pady=(0, 10), sticky="w")
        self.cb_type.set("")

        # Source Type
        lbl_source_type = ttk.Label(right_half, text="Source Type:", style="CardLabel.TLabel")
        lbl_source_type.grid(row=2, column=0, pady=(5, 2), sticky="w")
        self.cb_source_type = ctk.CTkComboBox(right_half, font=("Segoe UI", 16), width=400)
        self.cb_source_type.grid(row=3, column=0, pady=(0, 10), sticky="w")
        self.cb_source_type.set("")

        # Ledger Category
        lbl_ledger = ttk.Label(right_half, text="Ledger Category:", style="CardLabel.TLabel")
        lbl_ledger.grid(row=4, column=0, pady=(5, 2), sticky="w")
        self.cb_ledger = ctk.CTkComboBox(right_half, font=("Segoe UI", 16), width=400)
        self.cb_ledger.grid(row=5, column=0, pady=(0, 10), sticky="w")
        self.cb_ledger.set("")

        # Button Toolbar Container
        btn_frame = ctk.CTkFrame(self.transaction_tab, fg_color="transparent")
        btn_frame.grid(row=1, column=0, sticky="ew", pady=(0, 10))

        self.btn_add = ctk.CTkButton(btn_frame, text="Add", fg_color=self.theme_colors["accent"], width=80, command=self.add_transaction)
        self.btn_add.pack(side="left", padx=(0, 10))

        self.btn_update = ctk.CTkButton(btn_frame, text="Update", fg_color=self.theme_colors["accent"], width=80, command=self.update_transaction)
        self.btn_update.pack(side="left", padx=10)

        self.btn_delete = ctk.CTkButton(btn_frame, text="Delete", fg_color=self.theme_colors["accent"], width=80, command=self.delete_transaction)
        self.btn_delete.pack(side="left", padx=10)

        self.btn_retrieve = ctk.CTkButton(btn_frame, text="Retrieve", fg_color=self.theme_colors["accent"], width=80, command=self.retrieve_transaction)
        self.btn_retrieve.pack(side="left", padx=10)

        self.btn_clear = ctk.CTkButton(btn_frame, text="Clear", fg_color=self.theme_colors["accent"], width=80, command=self.clear_form)
        self.btn_clear.pack(side="left", padx=10)

        self.btn_print_voucher = ctk.CTkButton(btn_frame, text="Print Voucher", fg_color=self.theme_colors["accent"], width=110, command=self.print_voucher, state="disabled")
        self.btn_print_voucher.pack(side="left", padx=10)
        # Data Grid Container (Treeview)
        grid_frame = ctk.CTkFrame(self.transaction_tab, fg_color="transparent")
        grid_frame.grid(row=2, column=0, sticky="nsew")
        grid_frame.rowconfigure(0, weight=1)
        grid_frame.columnconfigure(0, weight=1)

        columns = ("No.", "Date", "Voucher Number", "Type", "Source Type", "Source Ref", "Description", "Ledger", "Financial Report Category", "Amount")
        self.tree = ttk.Treeview(grid_frame, columns=columns, show="headings", selectmode="browse")
        self.tree.configure(displaycolumns=[c for c in columns if c not in ("No.", "Financial Report Category")])

        # Configure columns header and widths
        for col in columns:
            if col in ("Ledger", "Date", "Voucher Number", "Type", "Source Type", "Source Ref", "Description", "Financial Report Category"):
                self.tree.heading(col, text=col, anchor="w")
            elif col == "Amount":
                self.tree.heading(col, text="Amount (₦)", anchor="e")
            else:
                self.tree.heading(col, text=col)

            if col == "No.":
                self.tree.column(col, width=50, minwidth=50, anchor="center")
            elif col == "Description":
                self.tree.column(col, width=250, minwidth=150, anchor="w")
            elif col in ("Date", "Type"):
                self.tree.column(col, width=100, minwidth=80, anchor="w")
            elif col == "Amount":
                self.tree.column(col, width=100, minwidth=80, anchor="e")
            else:
                self.tree.column(col, width=120, minwidth=100, anchor="w")

        self.tree.grid(row=0, column=0, sticky="nsew")

        v_scroll = ttk.Scrollbar(grid_frame, orient="vertical", command=self.tree.yview)
        v_scroll.grid(row=0, column=1, sticky="ns")
        h_scroll = ttk.Scrollbar(grid_frame, orient="horizontal", command=self.tree.xview)
        h_scroll.grid(row=1, column=0, sticky="ew")

        self.tree.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)

        # Double-click binds to retrieve record
        self.tree.bind("<Double-1>", lambda event: self.retrieve_transaction())

        # Selection binds to update Print Voucher button state
        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)

        # Status Bar
        self.lbl_trans_status = ttk.Label(self.transaction_tab, text="Total Transactions: 0", font=("Segoe UI", 11, "italic"))
        self.lbl_trans_status.grid(row=3, column=0, sticky="e", pady=(0, 5), padx=10)

        # Populate combobox values initial load
        self.refresh_comboboxes()

    def populate_ledger_tab(self):
        """
        Sets up the interactive data entry form, button toolbar, and Treeview grid
        on the 'Ledger Budget' tab.
        """
        # Configure responsive layout weights for the tab
        self.ledger_tab.columnconfigure(0, weight=1)
        self.ledger_tab.rowconfigure(2, weight=1)  # Grid row expands

        # Scroll container wraps the top form
        scroll_container = ctk.CTkFrame(self.ledger_tab, fg_color="transparent")
        scroll_container.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        scroll_container.columnconfigure(0, weight=1)

        h_scroll_form = ttk.Scrollbar(scroll_container, orient="horizontal")
        h_scroll_form.pack(side="top", fill="x")

        form_canvas = tk.Canvas(scroll_container, bg=self.colors.get("bg", "#f8f9fa"), highlightthickness=0)
        form_canvas.pack(side="top", fill="x", expand=True)

        h_scroll_form.configure(command=form_canvas.xview)
        form_canvas.configure(xscrollcommand=h_scroll_form.set)

        card = ctk.CTkFrame(form_canvas, fg_color=self.theme_colors["card"], corner_radius=8, border_width=1, border_color=self.theme_colors["border"])
        form_window_id = form_canvas.create_window((0, 0), window=card, anchor="nw")

        def on_form_canvas_configure(event):
            req_width = card.winfo_reqwidth()
            if event.width > req_width:
                form_canvas.itemconfig(form_window_id, width=event.width)
            else:
                form_canvas.itemconfig(form_window_id, width=req_width)

        def on_card_configure(event):
            form_canvas.configure(scrollregion=form_canvas.bbox("all"))
            form_canvas.configure(height=card.winfo_reqheight())

        form_canvas.bind("<Configure>", on_form_canvas_configure)
        card.bind("<Configure>", on_card_configure)

        import calendar
        for col_idx in range(4):
            card.columnconfigure(col_idx, weight=0, minsize=140)

        # Add a dummy 5th column to absorb extra space when the window expands
        card.columnconfigure(4, weight=1)

        # Form Inputs Grid
        lbl_ledger = ttk.Label(card, text="Ledger Category:", style="CardLabel.TLabel")
        lbl_ledger.grid(row=0, column=0, padx=15, pady=(15, 2), sticky="w")
        self.cb_l_ledger = ctk.CTkComboBox(card, font=("Segoe UI", 13), width=250)
        self.cb_l_ledger.grid(row=1, column=0, padx=15, pady=(0, 15), sticky="ew")
        self.cb_l_ledger.set("<None>")
        self.cb_l_ledger.configure(command=lambda e: self.load_workbook_data())

        lbl_month = ttk.Label(card, text="Month:", style="CardLabel.TLabel")
        lbl_month.grid(row=0, column=1, padx=15, pady=(15, 2), sticky="w")
        self.cb_l_month = ctk.CTkComboBox(card, font=("Segoe UI", 13), values=list(calendar.month_name)[1:], width=120)
        self.cb_l_month.grid(row=1, column=1, padx=15, pady=(0, 15), sticky="ew")
        self.cb_l_month.set(list(calendar.month_name)[datetime.now().month])

        lbl_year = ttk.Label(card, text="Year:", style="CardLabel.TLabel")
        lbl_year.grid(row=0, column=2, padx=15, pady=(15, 2), sticky="w")
        self.ent_l_year = make_combo_spinner(card, datetime.now().year, [str(y) for y in range(datetime.now().year, datetime.now().year + 10)])
        self.ent_l_year.grid(row=1, column=2, padx=15, pady=(0, 15), sticky="w")

        lbl_amount = ttk.Label(card, text="Budget Amount:", style="CardLabel.TLabel")
        lbl_amount.grid(row=0, column=3, padx=15, pady=(15, 2), sticky="w")
        self.ent_l_amount = ctk.CTkEntry(card, font=("Segoe UI", 13), width=120)
        self.ent_l_amount.grid(row=1, column=3, padx=15, pady=(0, 15), sticky="w")
        self.ent_l_amount.bind("<FocusOut>", self.format_ledger_amount_field)

        # Button Toolbar Container
        btn_frame = ctk.CTkFrame(self.ledger_tab, fg_color="transparent")
        btn_frame.grid(row=1, column=0, sticky="ew", pady=(0, 10))

        self.btn_l_add = ctk.CTkButton(btn_frame, text="Add", fg_color=self.theme_colors["accent"], width=80, command=self.add_ledger_budget)
        self.btn_l_add.pack(side="left", padx=(0, 10))

        self.btn_l_update = ctk.CTkButton(btn_frame, text="Update", fg_color=self.theme_colors["accent"], width=80, command=self.update_ledger_budget)
        self.btn_l_update.pack(side="left", padx=10)

        self.btn_l_delete = ctk.CTkButton(btn_frame, text="Delete", fg_color=self.theme_colors["accent"], width=80, command=self.delete_ledger_budget)
        self.btn_l_delete.pack(side="left", padx=10)

        self.btn_l_retrieve = ctk.CTkButton(btn_frame, text="Retrieve", fg_color=self.theme_colors["accent"], width=80, command=self.retrieve_ledger_budget)
        self.btn_l_retrieve.pack(side="left", padx=10)

        self.btn_l_clear = ctk.CTkButton(btn_frame, text="Clear", fg_color=self.theme_colors["accent"], width=80, command=self.clear_ledger_form)
        self.btn_l_clear.pack(side="left", padx=10)

        # Data Grid Container (Treeview)
        grid_frame = ctk.CTkFrame(self.ledger_tab, fg_color="transparent")
        grid_frame.grid(row=2, column=0, sticky="nsew")
        grid_frame.rowconfigure(0, weight=1)
        grid_frame.columnconfigure(0, weight=1)

        columns = ("No.", "Ledger", "Month", "Year", "Amount")
        self.ledger_tree = ttk.Treeview(grid_frame, columns=columns, show="headings", selectmode="browse")

        # Configure columns header and widths
        for col in columns:
            if col in ("Ledger", "Month", "Year"):
                self.ledger_tree.heading(col, text=col, anchor="w")
            elif col == "Amount":
                self.ledger_tree.heading(col, text="Amount (₦)", anchor="e")
            else:
                self.ledger_tree.heading(col, text=col)
                
            if col == "No.":
                self.ledger_tree.column(col, width=50, minwidth=50, anchor="center")
            elif col == "Ledger":
                self.ledger_tree.column(col, width=200, minwidth=120)
            elif col in ("Month", "Year"):
                self.ledger_tree.column(col, width=120, minwidth=100, anchor="w")
            elif col == "Amount":
                self.ledger_tree.column(col, width=120, minwidth=100, anchor="e")
                
        # Hide "No." column
        self.ledger_tree.configure(displaycolumns=("Ledger", "Month", "Year", "Amount"))

        self.ledger_tree.grid(row=0, column=0, sticky="nsew")

        v_scroll = ttk.Scrollbar(grid_frame, orient="vertical", command=self.ledger_tree.yview)
        v_scroll.grid(row=0, column=1, sticky="ns")
        h_scroll = ttk.Scrollbar(grid_frame, orient="horizontal", command=self.ledger_tree.xview)
        h_scroll.grid(row=1, column=0, sticky="ew")

        self.ledger_tree.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)

        # Double-click binds to retrieve record
        self.ledger_tree.bind("<Double-1>", lambda event: self.retrieve_ledger_budget())

        # Populate combobox values initial load
        self.refresh_comboboxes()

    def populate_dashboard_tab(self):
        """Sets up the scrollable layout and dashboard cards on the Dashboard tab."""
        self.dashboard_tab.columnconfigure(0, weight=1)
        self.dashboard_tab.rowconfigure(1, weight=1)

        # 1. Top Control Toolbar (Refresh, Date selection, View Mode)
        controls_frame = ctk.CTkFrame(self.dashboard_tab, fg_color="transparent")
        controls_frame.grid(row=0, column=0, sticky="ew", padx=15, pady=10)


        months_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        years_list = [str(y) for y in range(datetime.now().year, datetime.now().year + 10)]

        lbl_from = ttk.Label(controls_frame, text="From:", style="TLabel")
        lbl_from.pack(side="left", padx=(0, 5))

        self.ent_dash_from_year = make_combo_spinner(controls_frame, datetime.now().year, years_list)
        self.ent_dash_from_year.pack(side="left", padx=2)

        self.cb_dash_from_month = ctk.CTkComboBox(controls_frame, values=months_names, font=("Segoe UI", 13), width=90)
        self.cb_dash_from_month.set(months_names[datetime.now().month - 1])
        self.cb_dash_from_month.pack(side="left", padx=(2, 10))

        lbl_to = ttk.Label(controls_frame, text="To:", style="TLabel")
        lbl_to.pack(side="left", padx=(5, 5))

        self.ent_dash_to_year = make_combo_spinner(controls_frame, datetime.now().year, years_list)
        self.ent_dash_to_year.pack(side="left", padx=2)

        self.cb_dash_to_month = ctk.CTkComboBox(controls_frame, values=months_names, font=("Segoe UI", 13), width=90)
        self.cb_dash_to_month.set(months_names[datetime.now().month - 1])
        self.cb_dash_to_month.pack(side="left", padx=(2, 10))

        self.cb_dash_from_month.configure(command=lambda e: self.render_dashboard())
        self.ent_dash_from_year.configure(command=lambda e: self.render_dashboard())
        self.cb_dash_to_month.configure(command=lambda e: self.render_dashboard())
        self.ent_dash_to_year.configure(command=lambda e: self.render_dashboard())

        self.btn_dash_refresh = ctk.CTkButton(controls_frame, text="Refresh Dashboard", text_color="#ffffff", fg_color=self.theme_colors["accent"], command=self.refresh_dashboard)
        self.btn_dash_refresh.pack(side="right", padx=5)

        # 2. Scrollable Canvas
        self.dash_canvas = tk.Canvas(self.dashboard_tab, bg=self.colors["bg"], bd=0, highlightthickness=0)
        self.dash_scrollbar = ttk.Scrollbar(self.dashboard_tab, orient="vertical", command=self.dash_canvas.yview)
        self.dash_scroll_content = ctk.CTkFrame(self.dash_canvas, fg_color="transparent")

        self.dash_scroll_content.bind(
            "<Configure>",
            lambda e: self.dash_canvas.configure(scrollregion=self.dash_canvas.bbox("all"))
        )
        self.dash_canvas_window = self.dash_canvas.create_window((0, 0), window=self.dash_scroll_content, anchor="nw")

        def on_canvas_configure(event):
            self.dash_canvas.itemconfig(self.dash_canvas_window, width=event.width)
        self.dash_canvas.bind("<Configure>", on_canvas_configure)

        self.dash_canvas.configure(yscrollcommand=self.dash_scrollbar.set)

        self.dash_canvas.grid(row=1, column=0, sticky="nsew")
        self.dash_scrollbar.grid(row=1, column=1, sticky="ns")

        # 3. Card Sections Grid
        self.dash_scroll_content.columnconfigure(0, weight=1, uniform="dash_col")
        self.dash_scroll_content.columnconfigure(1, weight=1, uniform="dash_col")

        # Type Summary Card
        self.card_type = ctk.CTkFrame(self.dash_scroll_content, fg_color=self.theme_colors["card"], corner_radius=8, border_width=1, border_color=self.theme_colors["border"])
        self.card_type.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")
        self.card_type.columnconfigure(0, weight=1)
        self.card_type.rowconfigure(1, weight=1)

        hdr_type = ctk.CTkFrame(self.card_type, fg_color="transparent")
        hdr_type.grid(row=0, column=0, sticky="ew", pady=(10, 10), padx=15)
        hdr_type.columnconfigure(0, weight=1)
        lbl_type_title = ttk.Label(hdr_type, text="Monthly Summary by Transaction Type", style="Header.TLabel", font=("Segoe UI", 13, "bold"))
        lbl_type_title.grid(row=0, column=0, sticky="w")
        btn_type_print = ctk.CTkButton(hdr_type, text="Print Report", width=90, text_color="#ffffff", fg_color=self.theme_colors["accent"], command=lambda: self.export_dashboard_report("type"))
        btn_type_print.grid(row=0, column=1, sticky="e")

        self.content_dash_type = ctk.CTkFrame(self.card_type, fg_color="transparent")
        self.content_dash_type.grid(row=1, column=0, sticky="nsew", padx=15, pady=(0, 15))

        # Source Summary Card
        self.card_source = ctk.CTkFrame(self.dash_scroll_content, fg_color=self.theme_colors["card"], corner_radius=8, border_width=1, border_color=self.theme_colors["border"])
        self.card_source.grid(row=0, column=1, padx=10, pady=10, sticky="nsew")
        self.card_source.columnconfigure(0, weight=1)
        self.card_source.rowconfigure(1, weight=1)

        hdr_source = ctk.CTkFrame(self.card_source, fg_color="transparent")
        hdr_source.grid(row=0, column=0, sticky="ew", pady=(10, 10), padx=15)
        hdr_source.columnconfigure(0, weight=1)
        lbl_source_title = ttk.Label(hdr_source, text="Monthly Summary by Source Type", style="Header.TLabel", font=("Segoe UI", 13, "bold"))
        lbl_source_title.grid(row=0, column=0, sticky="w")
        btn_source_print = ctk.CTkButton(hdr_source, text="Print Report", width=90, text_color="#ffffff", fg_color=self.theme_colors["accent"], command=lambda: self.export_dashboard_report("source"))
        btn_source_print.grid(row=0, column=1, sticky="e")

        self.content_dash_source = ctk.CTkFrame(self.card_source, fg_color="transparent")
        self.content_dash_source.grid(row=1, column=0, sticky="nsew", padx=15, pady=(0, 15))
        # Budget Performance Card
        self.card_budget = ctk.CTkFrame(self.dash_scroll_content, fg_color=self.theme_colors["card"], corner_radius=8, border_width=1, border_color=self.theme_colors["border"])
        self.card_budget.grid(row=1, column=0, columnspan=2, padx=10, pady=10, sticky="nsew")
        self.card_budget.columnconfigure(0, weight=1)
        self.card_budget.rowconfigure(1, weight=1)
        
        hdr_budget = ctk.CTkFrame(self.card_budget, fg_color="transparent")
        hdr_budget.grid(row=0, column=0, sticky="ew", pady=(10, 10), padx=15)
        hdr_budget.columnconfigure(0, weight=1)
        lbl_budget_title = ttk.Label(hdr_budget, text="Budget Performance Report", style="Header.TLabel", font=("Segoe UI", 13, "bold"))
        lbl_budget_title.grid(row=0, column=0, sticky="w")
        
        cats = ["All Categories"] + self.config_data.get("fin_report_categories", [])
        self.list_budget_category = CheckboxDropdown(hdr_budget, values=cats, width=250, command=lambda: self.render_budget_performance())
        self.list_budget_category.grid(row=0, column=2, sticky="e", padx=(0, 10))
        
        self.chk_budget_color_var = ctk.StringVar(value="on")
        self.chk_budget_color = ctk.CTkCheckBox(hdr_budget, text="Highlight Variance", variable=self.chk_budget_color_var, onvalue="on", offvalue="off", font=("Segoe UI", 12), command=lambda: self.render_budget_performance())
        self.chk_budget_color.grid(row=0, column=3, sticky="e", padx=(0, 10))
        
        btn_budget_print = ctk.CTkButton(hdr_budget, text="Print Report", width=90, text_color="#ffffff", fg_color=self.theme_colors["accent"], command=lambda: self.export_dashboard_report("budget"))
        btn_budget_print.grid(row=0, column=4, sticky="e")
        
        self.content_dash_budget = ctk.CTkFrame(self.card_budget, fg_color="transparent")
        self.content_dash_budget.grid(row=1, column=0, sticky="nsew", padx=15, pady=(0, 15))

    def load_dashboard_data(self):
        """Loads and aggregates data from the current workbook for the dashboard."""
        self.dash_transaction_data = []
        self.dash_budget_data = []
        self.dash_years = set()
        
        if not self.current_workbook_path:
            return
            
        try:
            wb = openpyxl.load_workbook(self.current_workbook_path, data_only=True)
            
            # Load transaction entries
            if "TransactionEntries" in wb.sheetnames:
                ws = wb["TransactionEntries"]
                for r in range(2, ws.max_row + 1):
                    row_vals = [ws.cell(row=r, column=c).value for c in range(1, 10)]
                    if all(v is None for v in row_vals):
                        continue
                    
                    # Parse Date
                    date_val = row_vals[0]
                    dt = None
                    if isinstance(date_val, datetime):
                        dt = date_val.date()
                    elif date_val is not None:
                        try:
                            dt = datetime.strptime(str(date_val).strip(), "%Y-%m-%d").date()
                        except ValueError:
                            pass
                            
                    # Parse Amount
                    amt = 0.0
                    if row_vals[8] is not None:
                        try:
                            amt = float(row_vals[8])
                        except ValueError:
                            pass
                            
                    t_type = str(row_vals[2] if row_vals[2] is not None else "").strip()
                    s_type = str(row_vals[3] if row_vals[3] is not None else "").strip()
                    ledger = str(row_vals[6] if row_vals[6] is not None else "").strip()
                    
                    if dt:
                        self.dash_years.add(dt.year)
                        self.dash_transaction_data.append({
                            "year": dt.year,
                            "month": dt.month,
                            "type": t_type,
                            "source_type": s_type,
                            "ledger": ledger,
                            "amount": amt
                        })
                        
            # Load ledger budgets
            if "LedgerBudget" in wb.sheetnames:
                ws = wb["LedgerBudget"]
                for r in range(2, ws.max_row + 1):
                    row_vals = [ws.cell(row=r, column=c).value for c in range(1, 5)]
                    if all(v is None for v in row_vals):
                        continue
                    
                    ledger = str(row_vals[0] if row_vals[0] is not None else "").strip()
                    
                    month_str = str(row_vals[1]).strip() if row_vals[1] is not None else "1"
                    try:
                        month = int(month_str)
                    except ValueError:
                        m_lower = month_str.lower()
                        months_map = {"january":1, "jan":1, "february":2, "feb":2, "march":3, "mar":3, 
                                      "april":4, "apr":4, "may":5, "june":6, "jun":6, "july":7, "jul":7, 
                                      "august":8, "aug":8, "september":9, "sep":9, "october":10, "oct":10, 
                                      "november":11, "nov":11, "december":12, "dec":12}
                        month = months_map.get(m_lower, 1)
                        
                    try:
                        year = int(row_vals[2]) if row_vals[2] is not None else 2026
                    except ValueError:
                        year = 2026
                        
                    try:
                        amt = float(row_vals[3]) if row_vals[3] is not None else 0.0
                    except ValueError:
                        amt = 0.0
                        
                    self.dash_years.add(year)
                    self.dash_budget_data.append({
                        "ledger": ledger,
                        "month": month,
                        "year": year,
                        "amount": amt
                    })
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load dashboard data:\n{e}", parent=self)

    def refresh_dashboard(self):
        """Loads data from the active workbook and redraws all dashboard panels."""
        self.load_dashboard_data()
        self.render_dashboard()

    def render_dashboard(self):
        """Redraws all three dashboard panels based on control configurations."""
        if not self.current_workbook_path:
            return
            
        months_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        
        try:
            f_month = months_names.index(self.cb_dash_from_month.get()) + 1
            f_year = int(self.ent_dash_from_year.get())
        except (ValueError, AttributeError):
            f_month = 1
            f_year = datetime.now().year
            
        try:
            t_month = months_names.index(self.cb_dash_to_month.get()) + 1
            t_year = int(self.ent_dash_to_year.get())
        except (ValueError, AttributeError):
            t_month = datetime.now().month
            t_year = datetime.now().year
            
        # Ensure 'from' is <= 'to'
        if f_year > t_year or (f_year == t_year and f_month > t_month):
            self.card_type.grid_remove()
            self.card_source.grid_remove()
            self.card_budget.grid_remove()
            from tkinter import messagebox
            messagebox.showerror("Invalid Period", "The 'From' period cannot be later than the 'To' period.", parent=self)
            return
            
        # Generate range of (year, month) tuples
        period_range = []
        cy, cm = f_year, f_month
        while cy < t_year or (cy == t_year and cm <= t_month):
            period_range.append((cy, cm))
            cm += 1
            if cm > 12:
                cm = 1
                cy += 1
                
        # Ensure all cards are gridded in their default multi-table layout
        self.card_type.grid(row=0, column=0, columnspan=1, rowspan=1, padx=10, pady=10, sticky="nsew")
        self.card_source.grid(row=0, column=1, columnspan=1, rowspan=1, padx=10, pady=10, sticky="nsew")
        self.card_budget.grid(row=1, column=0, columnspan=2, rowspan=1, padx=10, pady=10, sticky="nsew")
        
        self.dash_current_period_range = period_range
        
        self.render_type_summary(period_range)
        self.render_source_summary(period_range)
        self.render_budget_performance(period_range)

    def render_type_summary(self, period_range):
        # Clear panel content
        for child in self.content_dash_type.winfo_children():
            child.destroy()
            
        types = self.config_data.get("types", ["Payment", "Receipt", "Journal"])
        
        monthly_data = {ym: {t: 0.0 for t in types} for ym in period_range}
        for t in self.dash_transaction_data:
            ym = (t["year"], t["month"])
            if ym in monthly_data and t["type"] in types:
                monthly_data[ym][t["type"]] += t["amount"]
                
        months_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        
        if True:
            # Table View using Treeview
            columns = ("Period",) + tuple(types)
            
            v_scroll = ttk.Scrollbar(self.content_dash_type, orient="vertical")
            v_scroll.pack(side="right", fill="y")
            
            tree = ttk.Treeview(self.content_dash_type, columns=columns, show="headings", height=8, yscrollcommand=v_scroll.set)
            tree_totals = ttk.Treeview(self.content_dash_type, columns=columns, show="", height=1)
            v_scroll.configure(command=tree.yview)
            
            tree.pack(side="top", fill="both", expand=True)
            tree_totals.pack(side="bottom", fill="x")
            
            tree.heading("Period", text="Period", anchor="w")
            tree.column("Period", width=100, anchor="w")
            tree_totals.column("Period", width=100, anchor="w")
            for t in types:
                tree.heading(t, text=t, anchor="e")
                tree.column(t, width=110, anchor="e")
                tree_totals.column(t, width=110, anchor="e")
                
            column_totals = {t: 0.0 for t in types}
            
            theme = self.config_data.get("theme", "light")
            bg_even = "#f9fafb" if theme == "light" else "#1e1e1e"
            bg_odd = "#ffffff" if theme == "light" else "#121212"
            bg_total = "#e5e7eb" if theme == "light" else "#27272a"
            
            tree.tag_configure("evenrow", background=bg_even)
            tree.tag_configure("oddrow", background=bg_odd)
            tree_totals.tag_configure("totalrow", font=("Segoe UI", 11, "bold"), background=bg_total)
            
            for i, ym in enumerate(period_range):
                y, m = ym
                period_label = f"{months_names[m-1]}-{str(y)[-2:]}"
                row_vals = [period_label]
                for t in types:
                    amt = monthly_data[ym][t]
                    row_vals.append(f"₦{amt:,.2f}")
                    column_totals[t] += amt
                    
                tags = ("evenrow",) if i % 2 == 0 else ("oddrow",)
                tree.insert("", tk.END, values=row_vals, tags=tags)
                
            # Add final row for column totals in the frozen treeview
            total_row_vals = ["TOTAL"]
            for t in types:
                total_row_vals.append(f"₦{column_totals[t]:,.2f}")
            tree_totals.insert("", tk.END, values=total_row_vals, tags=("totalrow",))
            
            def sync_columns_type(event):
                for col in columns:
                    tree_totals.column(col, width=tree.column(col, "width"))
            tree.bind("<B1-Motion>", sync_columns_type)
            tree.bind("<ButtonRelease-1>", sync_columns_type)

    def render_source_summary(self, period_range):
        # Clear panel content
        for child in self.content_dash_source.winfo_children():
            child.destroy()
            
        source_types = self.config_data.get("source_types", ["Bank", "Cash", "Credit Card"])
        
        monthly_data = {ym: {s: 0.0 for s in source_types} for ym in period_range}
        for t in self.dash_transaction_data:
            ym = (t["year"], t["month"])
            if ym in monthly_data and t["source_type"] in source_types:
                monthly_data[ym][t["source_type"]] += t["amount"]
                
        months_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        
        if True:
            # Table View using Treeview
            columns = ("Period",) + tuple(source_types)
            
            v_scroll = ttk.Scrollbar(self.content_dash_source, orient="vertical")
            v_scroll.pack(side="right", fill="y")
            
            tree = ttk.Treeview(self.content_dash_source, columns=columns, show="headings", height=8, yscrollcommand=v_scroll.set)
            tree_totals = ttk.Treeview(self.content_dash_source, columns=columns, show="", height=1)
            v_scroll.configure(command=tree.yview)
            
            tree.pack(side="top", fill="both", expand=True)
            tree_totals.pack(side="bottom", fill="x")
            
            tree.heading("Period", text="Period", anchor="w")
            tree.column("Period", width=100, anchor="w")
            tree_totals.column("Period", width=100, anchor="w")
            for t in source_types:
                tree.heading(t, text=t, anchor="e")
                tree.column(t, width=110, anchor="e")
                tree_totals.column(t, width=110, anchor="e")
                
            column_totals = {t: 0.0 for t in source_types}
            
            theme = self.config_data.get("theme", "light")
            bg_even = "#f9fafb" if theme == "light" else "#1e1e1e"
            bg_odd = "#ffffff" if theme == "light" else "#121212"
            bg_total = "#e5e7eb" if theme == "light" else "#27272a"
            
            tree.tag_configure("evenrow", background=bg_even)
            tree.tag_configure("oddrow", background=bg_odd)
            tree_totals.tag_configure("totalrow", font=("Segoe UI", 11, "bold"), background=bg_total)
            
            for i, ym in enumerate(period_range):
                y, m = ym
                period_label = f"{months_names[m-1]}-{str(y)[-2:]}"
                row_vals = [period_label]
                for t in source_types:
                    amt = monthly_data[ym][t]
                    row_vals.append(f"₦{amt:,.2f}")
                    column_totals[t] += amt
                    
                tags = ("evenrow",) if i % 2 == 0 else ("oddrow",)
                tree.insert("", tk.END, values=row_vals, tags=tags)
                
            # Add final row for column totals in the frozen treeview
            total_row_vals = ["TOTAL"]
            for t in source_types:
                total_row_vals.append(f"₦{column_totals[t]:,.2f}")
            tree_totals.insert("", tk.END, values=total_row_vals, tags=("totalrow",))
            
            def sync_columns_source(event):
                for col in columns:
                    tree_totals.column(col, width=tree.column(col, "width"))
            tree.bind("<B1-Motion>", sync_columns_source)
            tree.bind("<ButtonRelease-1>", sync_columns_source)

    def render_budget_performance(self, period_range=None):
        if period_range is None:
            period_range = getattr(self, "dash_current_period_range", [])
            
        # Clear panel content
        for child in self.content_dash_budget.winfo_children():
            child.destroy()
            
        all_ledgers = self.config_data.get("ledgers", ["Sales", "Purchases", "Expenses", "Income"])
        dropdown = getattr(self, "list_budget_category", None)
        selected_categories = dropdown.get() if dropdown else []
            
        if not selected_categories or "All Categories" in selected_categories:
            ledgers = all_ledgers
            filter_text = "All Categories"
        else:
            mapping = self.config_data.get("ledger_category_mapping", {})
            ledgers = [l for l in all_ledgers if mapping.get(l, "Uncategorized") in selected_categories]
            filter_text = ", ".join(selected_categories)
            
        lbl_criteria = ctk.CTkLabel(self.content_dash_budget, text=f"Filtering By: {filter_text}", font=("Segoe UI", 11, "italic"), text_color="gray")
        lbl_criteria.pack(side="top", anchor="w", pady=(0, 5))
        
        budget_data = {ym: {l: 0.0 for l in ledgers} for ym in period_range}
        for b in self.dash_budget_data:
            ym = (b["year"], b["month"])
            if ym in budget_data and b["ledger"] in ledgers:
                budget_data[ym][b["ledger"]] += b["amount"]
                
        actual_data = {ym: {l: 0.0 for l in ledgers} for ym in period_range}
        for t in self.dash_transaction_data:
            ym = (t["year"], t["month"])
            if ym in actual_data and t["ledger"] in ledgers:
                actual_data[ym][t["ledger"]] += t["amount"]
                
        months_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        
        if True:
            # Table View using Treeview
            columns = ("Ledger", "Budget", "Actual", "Variance", "Variance %")
            
            v_scroll = ttk.Scrollbar(self.content_dash_budget, orient="vertical")
            v_scroll.pack(side="right", fill="y")
            
            tree = ttk.Treeview(self.content_dash_budget, columns=columns, show="headings", height=10, yscrollcommand=v_scroll.set)
            tree_totals = ttk.Treeview(self.content_dash_budget, columns=columns, show="", height=1)
            v_scroll.configure(command=tree.yview)
            
            tree.pack(side="top", fill="both", expand=True)
            tree_totals.pack(side="bottom", fill="x")
            
            tree.heading("Ledger", text="Ledger Category", anchor="w")
            tree.column("Ledger", width=120, anchor="w")
            tree_totals.column("Ledger", width=120, anchor="w")
            tree.heading("Budget", text="Budget", anchor="e")
            tree.column("Budget", width=90, anchor="e")
            tree_totals.column("Budget", width=90, anchor="e")
            tree.heading("Actual", text="Actual", anchor="e")
            tree.column("Actual", width=90, anchor="e")
            tree_totals.column("Actual", width=90, anchor="e")
            tree.heading("Variance", text="Variance", anchor="e")
            tree.column("Variance", width=90, anchor="e")
            tree_totals.column("Variance", width=90, anchor="e")
            tree.heading("Variance %", text="Variance %", anchor="e")
            tree.column("Variance %", width=80, anchor="e")
            tree_totals.column("Variance %", width=80, anchor="e")
            
            theme = self.config_data.get("theme", "light")
            bg_even = "#f9fafb" if theme == "light" else "#1e1e1e"
            bg_odd = "#ffffff" if theme == "light" else "#121212"
            bg_total = "#e5e7eb" if theme == "light" else "#27272a"
            green_color = "#047857" if theme == "light" else "#34D399"
            red_color = "#B91C1C" if theme == "light" else "#F87171"
            normal_color = "#1A2530" if theme == "light" else "#F1F5F9"

            use_color = getattr(self, "chk_budget_color_var", None) and self.chk_budget_color_var.get() == "on"
            pos_color = green_color if use_color else normal_color
            neg_color = red_color if use_color else normal_color

            tree.tag_configure("pos_even", background=bg_even, foreground=pos_color)
            tree.tag_configure("pos_odd", background=bg_odd, foreground=pos_color)
            tree.tag_configure("neg_even", background=bg_even, foreground=neg_color)
            tree.tag_configure("neg_odd", background=bg_odd, foreground=neg_color)
            tree.tag_configure("zero_even", background=bg_even, foreground=normal_color)
            tree.tag_configure("zero_odd", background=bg_odd, foreground=normal_color)

            has_rows = False
            total_b = 0.0
            total_a = 0.0
            
            row_idx = 0
            for l in ledgers:
                b_amt = sum(budget_data[ym][l] for ym in period_range)
                a_amt = sum(actual_data[ym][l] for ym in period_range)
                
                if b_amt == 0 and a_amt == 0:
                    continue
                    
                total_b += b_amt
                total_a += a_amt
                
                var_amt = a_amt - b_amt
                var_pct = (var_amt / b_amt * 100.0) if b_amt != 0.0 else 0.0
                
                var_pct_str = f"{var_pct:+.1f}%" if b_amt != 0.0 else "N/A"
                var_amt_str = f"₦{var_amt:+,.2f}" if var_amt != 0.0 else "₦0.00"
                
                if var_amt > 0:
                    tag_prefix = "neg_"
                elif var_amt < 0:
                    tag_prefix = "pos_"
                else:
                    tag_prefix = "zero_"
                tag_suffix = "even" if row_idx % 2 == 0 else "odd"
                tags = (tag_prefix + tag_suffix,)
                
                tree.insert("", tk.END, values=(
                    l,
                    f"₦{b_amt:,.2f}",
                    f"₦{a_amt:,.2f}",
                    var_amt_str,
                    var_pct_str
                ), tags=tags)
                has_rows = True
                row_idx += 1
                
            if not has_rows:
                tree.insert("", tk.END, values=("No budget or transaction entries found for this period.", "", "", "", ""))
                
            # Render Totals Row
            total_var = total_a - total_b
            total_var_pct = (total_var / total_b * 100.0) if total_b != 0.0 else 0.0
            total_var_pct_str = f"{total_var_pct:+.1f}%" if total_b != 0.0 else "N/A"
            total_var_str = f"₦{total_var:+,.2f}" if total_var != 0.0 else "₦0.00"
            
            if use_color:
                total_fg = red_color if total_var > 0 else (green_color if total_var < 0 else normal_color)
            else:
                total_fg = normal_color
            tree_totals.tag_configure("totalrow", font=("Segoe UI", 11, "bold"), background=bg_total, foreground=total_fg)
            
            tree_totals.insert("", tk.END, values=(
                "TOTAL",
                f"₦{total_b:,.2f}",
                f"₦{total_a:,.2f}",
                total_var_str,
                total_var_pct_str
            ), tags=("totalrow",))
            
            def sync_columns_budget(event):
                for col in columns:
                    tree_totals.column(col, width=tree.column(col, "width"))
            tree.bind("<B1-Motion>", sync_columns_budget)
            tree.bind("<ButtonRelease-1>", sync_columns_budget)

    def export_dashboard_report(self, report_type):
        """Export the selected dashboard report to PDF using fpdf2."""
        try:
            from fpdf import FPDF
        except ImportError:
            messagebox.showerror("Export Error", "fpdf2 library not found. Please run 'pip install fpdf2'.", parent=self)
            return

        from tkinter import filedialog, messagebox
        
        months_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        
        # Determine active period range
        from_m = months_names.index(self.cb_dash_from_month.get()) + 1
        from_y = int(self.ent_dash_from_year.get())
        to_m = months_names.index(self.cb_dash_to_month.get()) + 1
        to_y = int(self.ent_dash_to_year.get())
        
        period_range = []
        cy, cm = from_y, from_m
        while (cy < to_y) or (cy == to_y and cm <= to_m):
            period_range.append((cy, cm))
            cm += 1
            if cm > 12:
                cm = 1
                cy += 1

        pdf = FPDF(orientation="P", unit="mm", format="A4")
        pdf.add_page()
        pdf.set_font("helvetica", style="B", size=16)
        
        subtitle = ""
        if report_type == "type":
            title = "Monthly Summary by Transaction Type"
            categories = self.config_data.get("types", [])
            data_dict = {ym: {t: 0.0 for t in categories} for ym in period_range}
            for t in self.dash_transaction_data:
                ym = (t["year"], t["month"])
                if ym in data_dict and t["type"] in categories:
                    data_dict[ym][t["type"]] += t["amount"]
                    
        elif report_type == "source":
            title = "Monthly Summary by Source Type"
            categories = self.config_data.get("source_types", [])
            data_dict = {ym: {t: 0.0 for t in categories} for ym in period_range}
            for t in self.dash_transaction_data:
                ym = (t["year"], t["month"])
                if ym in data_dict and t["source_type"] in categories:
                    data_dict[ym][t["source_type"]] += t["amount"]
                    
        elif report_type == "budget":
            title = "Budget Performance Report"
            all_ledgers = self.config_data.get("ledgers", [])
            dropdown = getattr(self, "list_budget_category", None)
            selected_categories = dropdown.get() if dropdown else []
                
            if selected_categories and "All Categories" not in selected_categories:
                mapping = self.config_data.get("ledger_category_mapping", {})
                categories = [l for l in all_ledgers if mapping.get(l, "Uncategorized") in selected_categories]
                subtitle = f"(Filtered By: {', '.join(selected_categories)})"
            else:
                categories = all_ledgers

            budget_data = {ym: {l: 0.0 for l in categories} for ym in period_range}
            for b in self.dash_budget_data:
                ym = (b["year"], b["month"])
                if ym in budget_data and b["ledger"] in categories:
                    budget_data[ym][b["ledger"]] += b["amount"]
                    
            actual_data = {ym: {l: 0.0 for l in categories} for ym in period_range}
            for t in self.dash_transaction_data:
                ym = (t["year"], t["month"])
                if ym in actual_data and t["ledger"] in categories:
                    actual_data[ym][t["ledger"]] += t["amount"]
        else:
            return

        import os
        company_info = self.config_data.get("company_info", {})
        c_name = company_info.get("name", "")
        c_addr = company_info.get("address", "")
        c_phone = company_info.get("phone", "")
        c_email = company_info.get("email", "")
        c_web = company_info.get("website", "")
        c_logo = company_info.get("logo", "")

        if c_logo and os.path.isfile(c_logo):
            try:
                # Portrait A4 width is 210.
                # Right edge is 200. Logo width = 30. x = 200 - 30 = 170.
                pdf.image(c_logo, x=170, y=5, w=30)
            except Exception:
                pass

        if c_name:
            pdf.set_font("helvetica", "B", 16)
            pdf.set_text_color(30, 58, 138)
            pdf.set_x(45)
            pdf.multi_cell(120, 8, c_name, align="C")

            pdf.set_font("helvetica", "", 10)
            pdf.set_text_color(75, 85, 99)
            if c_addr:
                pdf.set_x(45)
                pdf.multi_cell(120, 6, c_addr, align="C")
            if c_phone:
                pdf.set_x(45)
                pdf.multi_cell(120, 6, f"Tel: {c_phone}", align="C")
            if c_email:
                pdf.set_x(45)
                pdf.multi_cell(120, 6, f"Email: {c_email}", align="C")
            if c_web:
                pdf.set_x(45)
                pdf.multi_cell(120, 6, c_web, align="C")
            pdf.ln(5)

        pdf.set_font("helvetica", "B", 14)
        pdf.set_text_color(30, 58, 138)
        pdf.multi_cell(0, 10, title, new_x="LMARGIN", new_y="NEXT", align="C")
        
        if subtitle:
            pdf.set_font("helvetica", "", 10)
            pdf.multi_cell(0, 6, subtitle, new_x="LMARGIN", new_y="NEXT", align="C")
            
        pdf.set_font("helvetica", size=10)
        pdf.set_text_color(75, 85, 99)
        pdf.multi_cell(0, 6, f"Period: {self.cb_dash_from_month.get()} {from_y} to {self.cb_dash_to_month.get()} {to_y}", new_x="LMARGIN", new_y="NEXT", align="C")
        pdf.ln(5)
        
        font_family = "helvetica"
        naira_symbol = "NGN"
        font_bold_path = "C:\\Windows\\Fonts\\arialbd.ttf"
        if os.path.exists(font_bold_path):
            try:
                import warnings
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore")
                    pdf.add_font("Arial", "B", font_bold_path)
                font_family = "Arial"
                naira_symbol = "\u20a6"
            except Exception:
                pass
        
        pdf.ln(5)
        pdf.set_text_color(0, 0, 0)

        if report_type in ("type", "source"):
            headers = ["Period"] + [f"{c} ({naira_symbol})" for c in categories]
            col_widths = [30] + [(190 - 30) / max(1, len(categories))] * len(categories)
            
            pdf.set_font(font_family, style="B", size=9)
            x_start = pdf.get_x()
            y_start = pdf.get_y()
            max_y = y_start

            for i, h in enumerate(headers):
                pdf.set_xy(x_start + sum(col_widths[:i]) + 1, y_start + 1)
                pdf.multi_cell(col_widths[i] - 2, 4, h, border=0, align="R" if i > 0 else "L")
                if pdf.get_y() + 1 > max_y:
                    max_y = pdf.get_y() + 1

            row_height = max(8, max_y - y_start)

            for i, h in enumerate(headers):
                pdf.set_xy(x_start + sum(col_widths[:i]), y_start)
                pdf.cell(col_widths[i], row_height, "", border=1)

            pdf.set_y(y_start + row_height)
            
            pdf.set_font("helvetica", size=9)
            totals = {c: 0.0 for c in categories}
            months_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
            for ym in period_range:
                pdf.cell(col_widths[0], 8, f"{months_names[ym[1]-1]}-{ym[0]}", border=1)
                for i, c in enumerate(categories):
                    val = data_dict[ym][c]
                    totals[c] += val
                    pdf.cell(col_widths[i+1], 8, f"{val:,.2f}", border=1, align="R")
                pdf.ln()
                
            pdf.set_font(font_family, style="B", size=9)
            pdf.cell(col_widths[0], 8, "TOTAL", border=1)
            for i, c in enumerate(categories):
                pdf.cell(col_widths[i+1], 8, f"{totals[c]:,.2f}", border=1, align="R")
            pdf.ln()
            
        elif report_type == "budget":
            headers = ["Ledger Category", f"Budget ({naira_symbol})", f"Actual ({naira_symbol})", f"Variance ({naira_symbol})", "Variance %"]
            col_widths = [50, 35, 35, 35, 35]
            
            pdf.set_font(font_family, style="B", size=9)
            for i, h in enumerate(headers):
                pdf.cell(col_widths[i], 8, h, border=1, align="R" if i > 0 else "L")
            pdf.ln()
            
            pdf.set_font("helvetica", size=9)
            total_b = 0.0
            total_a = 0.0
            
            print_color = getattr(self, "chk_budget_color_var", None)
            use_color = True if not print_color else (print_color.get() == "on")
            
            for c in categories:
                b_amt = sum(budget_data[ym][c] for ym in period_range)
                a_amt = sum(actual_data[ym][c] for ym in period_range)
                if b_amt == 0 and a_amt == 0:
                    continue
                    
                total_b += b_amt
                total_a += a_amt
                var_amt = a_amt - b_amt
                var_pct = (var_amt / b_amt * 100.0) if b_amt != 0.0 else 0.0
                var_pct_str = f"{var_pct:+.1f}%" if b_amt != 0.0 else "N/A"
                var_amt_str = f"{var_amt:+,.2f}"
                
                if use_color:
                    if var_amt > 0:
                        pdf.set_text_color(185, 28, 28)
                    elif var_amt < 0:
                        pdf.set_text_color(4, 120, 87)
                    else:
                        pdf.set_text_color(0, 0, 0)
                else:
                    pdf.set_text_color(0, 0, 0)
                
                x_start = pdf.get_x()
                y_start = pdf.get_y()
                
                text_w = pdf.get_string_width(c)
                lines = int(text_w / (col_widths[0] - 2)) + 1
                row_height = max(8, lines * 4 + 4)
                
                if y_start + row_height > 275:
                    pdf.add_page()
                    y_start = pdf.get_y()
                    x_start = pdf.get_x()
                
                pdf.set_xy(x_start + 1, y_start + 2)
                pdf.multi_cell(col_widths[0] - 2, 4, c, border=0, align="L")
                
                actual_y = pdf.get_y()
                row_height = max(8, actual_y - y_start + 2)
                
                pdf.set_xy(x_start, y_start)
                pdf.cell(col_widths[0], row_height, "", border=1)
                
                pdf.set_xy(x_start + col_widths[0], y_start)
                pdf.cell(col_widths[1], row_height, f"{b_amt:,.2f}", border=1, align="R")
                pdf.cell(col_widths[2], row_height, f"{a_amt:,.2f}", border=1, align="R")
                pdf.cell(col_widths[3], row_height, var_amt_str, border=1, align="R")
                pdf.cell(col_widths[4], row_height, var_pct_str, border=1, align="R")
                
                pdf.set_xy(x_start, y_start + row_height)
                pdf.set_text_color(0, 0, 0)
                
            total_var = total_a - total_b
            total_var_pct = (total_var / total_b * 100.0) if total_b != 0.0 else 0.0
            total_var_pct_str = f"{total_var_pct:+.1f}%" if total_b != 0.0 else "N/A"
            
            pdf.set_font("helvetica", style="B", size=9)
            
            if use_color:
                if total_var > 0:
                    pdf.set_text_color(185, 28, 28)
                elif total_var < 0:
                    pdf.set_text_color(4, 120, 87)
                else:
                    pdf.set_text_color(0, 0, 0)
            else:
                pdf.set_text_color(0, 0, 0)
                
            pdf.cell(col_widths[0], 8, "TOTAL", border=1)
            pdf.cell(col_widths[1], 8, f"{total_b:,.2f}", border=1, align="R")
            pdf.cell(col_widths[2], 8, f"{total_a:,.2f}", border=1, align="R")
            pdf.cell(col_widths[3], 8, f"{total_var:+,.2f}", border=1, align="R")
            pdf.cell(col_widths[4], 8, total_var_pct_str, border=1, align="R")
            pdf.ln()
            pdf.set_text_color(0, 0, 0)

        import os
        dirs = self.config_data.get("directories", {})
        reports_dir = dirs.get("reports_directory", "").strip()
        if not reports_dir or not os.path.isdir(reports_dir):
            wb_dir = os.path.dirname(self.current_workbook_path)
            reports_dir = os.path.join(wb_dir, "reports")
            
        os.makedirs(reports_dir, exist_ok=True)
        
        full_title_for_filename = title + ("_Filtered" if subtitle else "")
        safe_title = full_title_for_filename.replace("\n", "_").replace(":", "_").replace(" ", "_").replace("(", "").replace(")", "")
        if len(safe_title) > 100:
            safe_title = safe_title[:100] + "..."
        filename = f"{safe_title}_{from_y}_{from_m}_to_{to_y}_{to_m}.pdf"
        pdf_path = os.path.join(reports_dir, filename)
        
        try:
            pdf.output(pdf_path)
            os.startfile(pdf_path)
        except PermissionError:
            messagebox.showerror("Permission Error", f"Could not save the PDF because '{filename}' is currently open in another application.\nPlease close it and try again.", parent=self)
        except Exception as e:
            messagebox.showerror("Print Error", f"An error occurred while printing the report:\n{e}", parent=self)

class PreferencesFrame(ctk.CTkFrame):
    def __init__(self, parent_container, app_instance):
        super().__init__(parent_container, fg_color="transparent")
        self.parent = app_instance
        self.config_data = self.parent.config_data.get("directories", {})
        self.company_info = self.parent.config_data.get("company_info", {})

        main_frame = ctk.CTkFrame(self, fg_color="transparent")
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)

        tabview = ctk.CTkTabview(
            main_frame, 
            segmented_button_font=("Segoe UI", 15, "bold"),
            segmented_button_selected_color=self.parent.theme_colors["accent"],
            segmented_button_selected_hover_color=self.parent.theme_colors["accent"]
        )
        tabview.pack(fill="both", expand=True, pady=(0, 15))

        tab_dirs = tabview.add("Directories")
        tab_comp = tabview.add("Company Profile")

        # --- Directories Tab ---
        ctk.CTkLabel(tab_dirs, text="Working Directory (New Workbooks):", font=("Segoe UI", 13, "bold")).pack(anchor="w", pady=(10, 5))
        frame1 = ctk.CTkFrame(tab_dirs, fg_color="transparent")
        frame1.pack(fill="x", pady=(0, 15))
        self.ent_work_dir = ctk.CTkEntry(frame1)
        self.ent_work_dir.pack(side="left", fill="x", expand=True, padx=(0, 10))
        self.ent_work_dir.insert(0, self.config_data.get("working_directory", ""))
        ctk.CTkButton(frame1, text="Browse...", width=90, text_color="#ffffff", fg_color=self.parent.theme_colors["accent"], command=lambda: self.browse_dir(self.ent_work_dir)).pack(side="right")

        ctk.CTkLabel(tab_dirs, text="Dashboard Reports Directory:", font=("Segoe UI", 13, "bold")).pack(anchor="w", pady=(0, 5))
        frame2 = ctk.CTkFrame(tab_dirs, fg_color="transparent")
        frame2.pack(fill="x", pady=(0, 15))
        self.ent_rep_dir = ctk.CTkEntry(frame2)
        self.ent_rep_dir.pack(side="left", fill="x", expand=True, padx=(0, 10))
        self.ent_rep_dir.insert(0, self.config_data.get("reports_directory", ""))
        ctk.CTkButton(frame2, text="Browse...", width=90, text_color="#ffffff", fg_color=self.parent.theme_colors["accent"], command=lambda: self.browse_dir(self.ent_rep_dir)).pack(side="right")

        ctk.CTkLabel(tab_dirs, text="Transactions Report Directory:", font=("Segoe UI", 13, "bold")).pack(anchor="w", pady=(0, 5))
        frame_trans_rep = ctk.CTkFrame(tab_dirs, fg_color="transparent")
        frame_trans_rep.pack(fill="x", pady=(0, 15))
        self.ent_trans_rep_dir = ctk.CTkEntry(frame_trans_rep)
        self.ent_trans_rep_dir.pack(side="left", fill="x", expand=True, padx=(0, 10))
        self.ent_trans_rep_dir.insert(0, self.config_data.get("transactions_report_directory", ""))
        ctk.CTkButton(frame_trans_rep, text="Browse...", width=90, text_color="#ffffff", fg_color=self.parent.theme_colors["accent"], command=lambda: self.browse_dir(self.ent_trans_rep_dir)).pack(side="right")

        ctk.CTkLabel(tab_dirs, text="Printed Vouchers Directory:", font=("Segoe UI", 13, "bold")).pack(anchor="w", pady=(0, 5))
        frame3 = ctk.CTkFrame(tab_dirs, fg_color="transparent")
        frame3.pack(fill="x", pady=(0, 20))
        self.ent_vouc_dir = ctk.CTkEntry(frame3)
        self.ent_vouc_dir.pack(side="left", fill="x", expand=True, padx=(0, 10))
        self.ent_vouc_dir.insert(0, self.config_data.get("vouchers_directory", ""))
        ctk.CTkButton(frame3, text="Browse...", width=90, text_color="#ffffff", fg_color=self.parent.theme_colors["accent"], command=lambda: self.browse_dir(self.ent_vouc_dir)).pack(side="right")

        # --- Company Profile Tab ---
        comp_container = ctk.CTkFrame(tab_comp, fg_color="transparent")
        comp_container.pack(fill="both", expand=True, padx=10, pady=10)

        ctk.CTkLabel(comp_container, text="Company Name:", font=("Segoe UI", 13, "bold")).grid(row=0, column=0, sticky="w", pady=(0, 10))
        self.ent_comp_name = ctk.CTkEntry(comp_container, width=500)
        self.ent_comp_name.grid(row=0, column=1, sticky="w", padx=10, pady=(0, 10))
        self.ent_comp_name.insert(0, self.company_info.get("name", ""))

        ctk.CTkLabel(comp_container, text="Company Address:", font=("Segoe UI", 13, "bold")).grid(row=1, column=0, sticky="nw", pady=(0, 10))
        self.ent_comp_addr = ctk.CTkTextbox(comp_container, width=500, height=80, font=("Segoe UI", 13), corner_radius=6, border_width=2)
        self.ent_comp_addr.grid(row=1, column=1, sticky="w", padx=10, pady=(0, 10))
        self.ent_comp_addr.insert("1.0", self.company_info.get("address", ""))

        ctk.CTkLabel(comp_container, text="Phone Number:", font=("Segoe UI", 13, "bold")).grid(row=2, column=0, sticky="w", pady=(0, 10))
        self.ent_comp_phone = ctk.CTkEntry(comp_container, width=500)
        self.ent_comp_phone.grid(row=2, column=1, sticky="w", padx=10, pady=(0, 10))
        self.ent_comp_phone.insert(0, self.company_info.get("phone", ""))

        ctk.CTkLabel(comp_container, text="Email Address:", font=("Segoe UI", 13, "bold")).grid(row=3, column=0, sticky="w", pady=(0, 10))
        self.ent_comp_email = ctk.CTkEntry(comp_container, width=500)
        self.ent_comp_email.grid(row=3, column=1, sticky="w", padx=10, pady=(0, 10))
        self.ent_comp_email.insert(0, self.company_info.get("email", ""))

        ctk.CTkLabel(comp_container, text="Website:", font=("Segoe UI", 13, "bold")).grid(row=4, column=0, sticky="w", pady=(0, 10))
        self.ent_comp_web = ctk.CTkEntry(comp_container, width=500)
        self.ent_comp_web.grid(row=4, column=1, sticky="w", padx=10, pady=(0, 10))
        self.ent_comp_web.insert(0, self.company_info.get("website", ""))

        ctk.CTkLabel(comp_container, text="Company Logo:", font=("Segoe UI", 13, "bold")).grid(row=5, column=0, sticky="w", pady=(0, 10))
        frame_logo = ctk.CTkFrame(comp_container, fg_color="transparent")
        frame_logo.grid(row=5, column=1, sticky="w", padx=10, pady=(0, 10))
        self.ent_comp_logo = ctk.CTkEntry(frame_logo, width=400)
        self.ent_comp_logo.pack(side="left", fill="x", expand=True)
        self.ent_comp_logo.insert(0, self.company_info.get("logo", ""))
        ctk.CTkButton(frame_logo, text="Browse...", width=90, text_color="#ffffff", fg_color=self.parent.theme_colors["accent"], command=self.browse_logo).pack(side="left", padx=(10, 0))
        
        self.lbl_logo_preview = ctk.CTkLabel(comp_container, text="")
        self.lbl_logo_preview.grid(row=0, column=2, rowspan=6, sticky="sw", padx=(40, 0), pady=(0, 10))
        self.update_logo_preview(self.company_info.get("logo", ""))

        # Buttons
        btn_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        btn_frame.pack(fill="x", pady=10)
        ctk.CTkButton(btn_frame, text="Save Preferences", text_color="#ffffff", fg_color=self.parent.theme_colors["accent"], command=self.save_preferences).pack(side="right")

    def browse_logo(self):
        from tkinter import filedialog
        file_path = filedialog.askopenfilename(parent=self, title="Select Logo", filetypes=[("Image Files", "*.png;*.jpg;*.jpeg;*.bmp")])
        if file_path:
            self.ent_comp_logo.delete(0, tk.END)
            self.ent_comp_logo.insert(0, file_path)
            self.update_logo_preview(file_path)
            
    def update_logo_preview(self, file_path):
        import os
        try:
            from PIL import Image
        except ImportError:
            self.lbl_logo_preview.configure(image="", text="[Pillow missing]")
            return
            
        if file_path and os.path.isfile(file_path):
            try:
                img = Image.open(file_path)
                img.thumbnail((150, 150))
                ctk_img = ctk.CTkImage(light_image=img, size=img.size)
                self.lbl_logo_preview.configure(image=ctk_img, text="")
                self.lbl_logo_preview.image = ctk_img
            except Exception:
                self.lbl_logo_preview.configure(image="", text="[Invalid Image]")
        else:
            self.lbl_logo_preview.configure(image="", text="[No Logo]")

    def browse_dir(self, entry_widget):
        from tkinter import filedialog
        directory = filedialog.askdirectory(parent=self, title="Select Directory")
        if directory:
            entry_widget.delete(0, tk.END)
            entry_widget.insert(0, directory)

    def save_preferences(self):
        import re
        email = self.ent_comp_email.get().strip()
        if email and not re.match(r"[^@]+@[^@]+\.[^@]+", email):
            messagebox.showwarning("Validation Error", "Please enter a valid email address.", parent=self)
            return
            
        website = self.ent_comp_web.get().strip()
        if website and not re.match(r"^(https?://)?(www\.)?[\w\.-]+\.\w+.*$", website):
            messagebox.showwarning("Validation Error", "Please enter a valid website URL.", parent=self)
            return

        self.parent.config_data["directories"] = {
            "working_directory": self.ent_work_dir.get().strip(),
            "reports_directory": self.ent_rep_dir.get().strip(),
            "transactions_report_directory": self.ent_trans_rep_dir.get().strip(),
            "vouchers_directory": self.ent_vouc_dir.get().strip()
        }
        self.parent.config_data["company_info"] = {
            "name": self.ent_comp_name.get().strip(),
            "address": self.ent_comp_addr.get("1.0", tk.END).strip(),
            "phone": self.ent_comp_phone.get().strip(),
            "email": email,
            "website": website,
            "logo": self.ent_comp_logo.get().strip()
        }
        self.parent.save_config()
        messagebox.showinfo("Success", "Preferences saved successfully.", parent=self)
        
        if hasattr(self.parent, '_previous_screen') and self.parent._previous_screen:
            self.parent._previous_screen()
        else:
            self.parent.hide_all_frames()
            if hasattr(self.parent, 'screen_title_label'):
                self.parent.screen_title_label.configure(text="")

class CustomInputDialog(ctk.CTkToplevel):
    def __init__(self, parent, title_text, prompt, initialvalue="", theme_colors=None):
        super().__init__(parent)
        self.title(title_text)
        self.result = None
        
        self.transient(parent)
        self.grab_set()
        
        # UI Elements
        self.lbl_prompt = ctk.CTkLabel(self, text=prompt, font=("Segoe UI", 14, "bold"))
        self.lbl_prompt.pack(pady=(20, 10), padx=20, anchor="w")
        
        self.entry = ctk.CTkEntry(self, font=("Segoe UI", 16), width=360, height=40)
        self.entry.pack(pady=10, padx=20)
        self.entry.insert(0, initialvalue)
        self.entry.focus()
        self.entry.bind("<Return>", self.on_ok)
        
        btn_frame = ctk.CTkFrame(self, fg_color="transparent")
        btn_frame.pack(pady=(10, 20), padx=20, fill="x")
        
        accent_color = theme_colors.get("accent", "#0078D7") if theme_colors else "#0078D7"
        
        self.btn_ok = ctk.CTkButton(btn_frame, text="OK", command=self.on_ok, width=100, fg_color=accent_color)
        self.btn_ok.pack(side="right", padx=(10, 0))
        
        self.btn_cancel = ctk.CTkButton(btn_frame, text="Cancel", command=self.on_cancel, width=100, fg_color="gray")
        self.btn_cancel.pack(side="right")
        
        # Center on parent after updating geometry
        self.update_idletasks()
        width = 420
        height = 220
        parent_x = parent.winfo_rootx()
        parent_y = parent.winfo_rooty()
        parent_width = parent.winfo_width()
        parent_height = parent.winfo_height()
        
        x = parent_x + (parent_width // 2) - (width // 2)
        y = parent_y + (parent_height // 2) - (height // 2)
        self.geometry(f"{width}x{height}+{x}+{y}")
        
        self.wait_window(self)

    def on_ok(self, event=None):
        self.result = self.entry.get()
        self.destroy()

    def on_cancel(self):
        self.result = None
        self.destroy()

class SetupListsFrame(ctk.CTkFrame):
    """
    Settings frame that allows dynamic management of Configuration Lists:
    Types, Source Types, and Ledgers. Save updates to config.json immediately.
    """
    def __init__(self, parent_container, app_instance):
        super().__init__(parent_container, fg_color="transparent")
        self.parent = app_instance
        self.colors = self.parent.colors

        # Main layout frame
        main_frame = ctk.CTkFrame(self, fg_color="transparent")
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)

        main_frame.rowconfigure(0, weight=1, uniform="row")
        main_frame.rowconfigure(1, weight=1, uniform="row")
        main_frame.columnconfigure(0, weight=1)

        top_section = ctk.CTkFrame(main_frame, fg_color="transparent")
        top_section.grid(row=0, column=0, sticky="nsew", pady=(0, 10))
        top_section.columnconfigure(0, weight=1, uniform="col")
        top_section.columnconfigure(1, weight=1, uniform="col")
        top_section.rowconfigure(0, weight=1)

        bottom_section = ctk.CTkFrame(main_frame, fg_color="transparent")
        bottom_section.grid(row=1, column=0, sticky="nsew")
        bottom_section.columnconfigure(0, weight=1, uniform="col")
        bottom_section.columnconfigure(1, weight=1, uniform="col")
        bottom_section.rowconfigure(0, weight=1)

        # Render columns for each data type
        self.create_column(top_section, 0, "Transaction Types", "types")
        self.create_column(top_section, 1, "Source Types", "source_types")
        self.create_ledger_mapping_column(bottom_section, 0, "Ledgers")
        self.create_column(bottom_section, 1, "Financial Report Category", "fin_report_categories")

    def create_column(self, parent_frame, col_idx, title, config_key):
        """Creates a stylized column containing a header, listbox, scrollbar, and add/remove buttons."""
        col_frame = ctk.CTkFrame(parent_frame, fg_color="transparent")
        col_frame.grid(row=0, column=col_idx, sticky="nsew", padx=10, pady=10)
        col_frame.rowconfigure(1, weight=1)
        col_frame.columnconfigure(0, weight=1)

        # Header
        lbl_header = ttk.Label(col_frame, text=title, font=("Segoe UI", 13, "bold"), anchor="center")
        lbl_header.grid(row=0, column=0, pady=(0, 10), sticky="ew")

        # Listbox/Scrollbar Container
        list_container = ctk.CTkFrame(col_frame, fg_color="transparent")
        list_container.grid(row=1, column=0, sticky="nsew", pady=(0, 10))
        list_container.rowconfigure(0, weight=1)
        list_container.columnconfigure(0, weight=1)

        # Convert Listbox to Treeview for better grid appearance
        columns = (title,)
        listbox = ttk.Treeview(list_container, columns=columns, show="headings", selectmode="browse")
        listbox.heading(title, text=title, anchor="w")
        listbox.column(title, anchor="w")
        listbox.grid(row=0, column=0, sticky="nsew")

        # Configure row colors
        theme = self.parent.config_data.get("theme", "light")
        even_bg = "#121212" if theme == "dark" else "#f9fafb"
        odd_bg  = "#000000" if theme == "dark" else "#ffffff"
        text_fg = "#f9fafb" if theme == "dark" else "#1f2937"
        listbox.tag_configure('evenrow', background=even_bg, foreground=text_fg)
        listbox.tag_configure('oddrow', background=odd_bg, foreground=text_fg)
        
        if not hasattr(self, 'list_trees'):
            self.list_trees = []
        self.list_trees.append(listbox)

        # Scrollbar
        scrollbar = ttk.Scrollbar(list_container, orient="vertical", command=listbox.yview)
        scrollbar.grid(row=0, column=1, sticky="ns")
        listbox.configure(yscrollcommand=scrollbar.set)
        
        # Load initial items
        items = sorted(self.parent.config_data.get(config_key, []))
        for idx, item in enumerate(items):
            tag = "evenrow" if idx % 2 == 0 else "oddrow"
            listbox.insert("", tk.END, values=(item,), tags=(tag,))
            
        # Button controls
        btn_frame = ctk.CTkFrame(col_frame, fg_color="transparent")
        btn_frame.grid(row=2, column=0, sticky="ew")
        btn_frame.columnconfigure(0, weight=1)
        btn_frame.columnconfigure(1, weight=1)
        btn_frame.columnconfigure(2, weight=1)
        
        btn_add = ctk.CTkButton(btn_frame, text="Add", text_color="#ffffff", fg_color=self.parent.theme_colors["accent"], command=lambda: self.add_item(listbox, config_key))
        btn_add.grid(row=0, column=0, padx=(0, 2), sticky="ew")
        
        btn_edit = ctk.CTkButton(btn_frame, text="Edit", text_color="#ffffff", fg_color=self.parent.theme_colors["accent"], command=lambda: self.edit_item(listbox, config_key))
        btn_edit.grid(row=0, column=1, padx=(2, 2), sticky="ew")
        
        btn_remove = ctk.CTkButton(btn_frame, text="Remove", text_color="#ffffff", fg_color=self.parent.theme_colors["accent"], command=lambda: self.remove_item(listbox, config_key))
        btn_remove.grid(row=0, column=2, padx=(2, 0), sticky="ew")

    def _get_tree_items(self, tree):
        return [tree.item(item_id, "values")[0] for item_id in tree.get_children()]

    def add_item(self, listbox, config_key):
        """Prompts for a new item, inserts it, and immediately updates the JSON config file."""
        display_name = config_key.replace('_', ' ').title().replace('Fin ', 'Financial ')
        dialog = CustomInputDialog(self.parent, "Add Item", f"Enter new value for {display_name}:", theme_colors=self.parent.theme_colors)
        new_item = dialog.result
        if not new_item:
            return
        
        new_item = new_item.strip()
        if not new_item:
            return
            
        current_items = self._get_tree_items(listbox)
        if new_item in current_items:
            messagebox.showwarning("Duplicate", "This item already exists.", parent=self)
            return
            
        current_items.append(new_item)
        current_items.sort()
        for item_id in listbox.get_children():
            listbox.delete(item_id)
        for idx, item in enumerate(current_items):
            tag = "evenrow" if idx % 2 == 0 else "oddrow"
            listbox.insert("", tk.END, values=(item,), tags=(tag,))
        self.parent.update_config_list(config_key, current_items)

    def edit_item(self, listbox, config_key):
        """Edits the selected item, updates the listbox, and writes to config.json."""
        selected_indices = listbox.selection()
        if not selected_indices:
            messagebox.showwarning("No Selection", "Please select an item to edit.", parent=self)
            return

        item_id = selected_indices[0]
        selected_val = listbox.item(item_id, "values")[0]

        display_name = config_key.replace('_', ' ').title().replace('Fin ', 'Financial ')
        dialog = CustomInputDialog(self.parent, "Edit Item", f"Edit value for {display_name}:", initialvalue=selected_val, theme_colors=self.parent.theme_colors)
        new_val = dialog.result
        if not new_val or new_val.strip() == selected_val:
            return

        new_val = new_val.strip()
        current_items = self._get_tree_items(listbox)
        if new_val in current_items:
            messagebox.showwarning("Duplicate", "This item already exists.", parent=self)
            return

        current_items[current_items.index(selected_val)] = new_val
        current_items.sort()
        for i_id in listbox.get_children():
            listbox.delete(i_id)
        for idx, item in enumerate(current_items):
            tag = "evenrow" if idx % 2 == 0 else "oddrow"
            listbox.insert("", tk.END, values=(item,), tags=(tag,))
        
        self.parent.update_config_list(config_key, current_items)

    def remove_item(self, listbox, config_key):
        """Removes the selected item, updates the listbox, and immediately writes to config.json."""
        selected_indices = listbox.selection()
        if not selected_indices:
            messagebox.showwarning("No Selection", "Please select an item to remove.", parent=self)
            return
            
        item_id = selected_indices[0]
        selected_val = listbox.item(item_id, "values")[0]
        
        confirm = messagebox.askyesno(
            "Confirm Delete", 
            f"Are you sure you want to remove '{selected_val}'?", 
            parent=self
        )
        if not confirm:
            return
            
        current_items = self._get_tree_items(listbox)
        current_items.remove(selected_val)
        for i_id in listbox.get_children():
            listbox.delete(i_id)
        for idx, item in enumerate(current_items):
            tag = "evenrow" if idx % 2 == 0 else "oddrow"
            listbox.insert("", tk.END, values=(item,), tags=(tag,))
        self.parent.update_config_list(config_key, current_items)

    def create_ledger_mapping_column(self, parent_frame, col_idx, title):
        col_frame = ctk.CTkFrame(parent_frame, fg_color="transparent")
        col_frame.grid(row=0, column=col_idx, sticky="nsew", padx=10, pady=10)
        col_frame.rowconfigure(1, weight=1)
        col_frame.columnconfigure(0, weight=1)

        # Header
        lbl_header = ttk.Label(col_frame, text=title, font=("Segoe UI", 13, "bold"), anchor="center")
        lbl_header.grid(row=0, column=0, pady=(0, 10), sticky="ew")

        # Treeview Container
        list_container = ctk.CTkFrame(col_frame, fg_color="transparent")
        list_container.grid(row=1, column=0, sticky="nsew", pady=(0, 10))
        list_container.rowconfigure(0, weight=1)
        list_container.columnconfigure(0, weight=1)

        columns = ("Ledger", "Category")
        tree = ttk.Treeview(list_container, columns=columns, show="headings", selectmode="browse")
        tree.heading("Ledger", text="Ledger Name", anchor="w")
        tree.heading("Category", text="Financial Report Category", anchor="w")
        tree.column("Ledger", width=120, anchor="w")
        tree.column("Category", width=120, anchor="w")
        tree.grid(row=0, column=0, sticky="nsew")

        scrollbar = ttk.Scrollbar(list_container, orient="vertical", command=tree.yview)
        scrollbar.grid(row=0, column=1, sticky="ns")
        tree.configure(yscrollcommand=scrollbar.set)
        
        if not hasattr(self, 'list_trees'):
            self.list_trees = []
        self.list_trees.append(tree)

        theme = self.parent.config_data.get("theme", "light")
        even_bg = "#121212" if theme == "dark" else "#f9fafb"
        odd_bg  = "#000000" if theme == "dark" else "#ffffff"
        text_fg = "#f9fafb" if theme == "dark" else "#1f2937"
        tree.tag_configure('evenrow', background=even_bg, foreground=text_fg)
        tree.tag_configure('oddrow', background=odd_bg, foreground=text_fg)

        # Load initial items
        ledgers = sorted(self.parent.config_data.get("ledgers", []))
        mapping = self.parent.config_data.get("ledger_category_mapping", {})
        for idx, ledger in enumerate(ledgers):
            cat = mapping.get(ledger, "")
            tag = "evenrow" if idx % 2 == 0 else "oddrow"
            tree.insert("", tk.END, values=(ledger, cat), tags=(tag,))

        # Button controls
        btn_frame = ctk.CTkFrame(col_frame, fg_color="transparent")
        btn_frame.grid(row=2, column=0, sticky="ew")
        btn_frame.columnconfigure(0, weight=1)
        btn_frame.columnconfigure(1, weight=1)
        btn_frame.columnconfigure(2, weight=1)

        btn_add = ctk.CTkButton(btn_frame, text="Add Ledger", text_color="#ffffff", fg_color=self.parent.theme_colors["accent"], command=lambda: self.add_ledger_item(tree))
        btn_add.grid(row=0, column=0, padx=(0, 2), sticky="ew")
        
        btn_edit = ctk.CTkButton(btn_frame, text="Edit Selected", text_color="#ffffff", fg_color=self.parent.theme_colors["accent"], command=lambda: self.edit_ledger_item(tree))
        btn_edit.grid(row=0, column=1, padx=(2, 2), sticky="ew")

        btn_remove = ctk.CTkButton(btn_frame, text="Remove Selected", text_color="#ffffff", fg_color=self.parent.theme_colors["accent"], command=lambda: self.remove_ledger_item(tree))
        btn_remove.grid(row=0, column=2, padx=(2, 0), sticky="ew")

    def add_ledger_item(self, tree):
        dialog = ctk.CTkToplevel(self)
        dialog.title("Add Ledger")
        dialog.transient(self)
        dialog.grab_set()
        
        main_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        ctk.CTkLabel(main_frame, text="Ledger Name:", font=("Segoe UI", 14, "bold")).pack(anchor="w", pady=(0, 5))
        ent_name = ctk.CTkEntry(main_frame, font=("Segoe UI", 16), width=360, height=40)
        ent_name.pack(fill="x", pady=(0, 15))
        
        ctk.CTkLabel(main_frame, text="Financial Report Category:", font=("Segoe UI", 14, "bold")).pack(anchor="w", pady=(0, 5))
        categories = sorted(self.parent.config_data.get("fin_report_categories", []))
        cb_category = ctk.CTkComboBox(main_frame, values=categories, state="readonly", font=("Segoe UI", 16), width=360, height=40)
        cb_category.set("")
        cb_category.pack(fill="x", pady=(0, 15))
        
        def on_cancel():
            dialog.destroy()
            
        def on_save():
            name = ent_name.get().strip()
            category = cb_category.get().strip()
            
            if not name:
                messagebox.showerror("Error", "Ledger Name is required.", parent=dialog)
                return
                
            ledgers = self.parent.config_data.get("ledgers", [])
            if name in ledgers:
                messagebox.showwarning("Duplicate", "This ledger already exists.", parent=dialog)
                return
                
            # Update data structures
            ledgers.append(name)
            mapping = self.parent.config_data.get("ledger_category_mapping", {})
            if category:
                mapping[name] = category
                
            self.parent.config_data["ledgers"] = ledgers
            self.parent.config_data["ledger_category_mapping"] = mapping
            self.parent.save_config()
            self.parent.refresh_comboboxes()
            
            tree.delete(*tree.get_children())
            sorted_ledgers = sorted(ledgers)
            for idx, l in enumerate(sorted_ledgers):
                c = mapping.get(l, "")
                tag = "evenrow" if idx % 2 == 0 else "oddrow"
                tree.insert("", tk.END, values=(l, c), tags=(tag,))
            dialog.destroy()
            
        btn_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        btn_frame.pack(fill="x", pady=(10, 0))
        
        accent_color = self.parent.theme_colors.get("accent", "#0078D7")
        
        btn_save = ctk.CTkButton(btn_frame, text="OK", text_color="#ffffff", fg_color=accent_color, width=100, command=on_save)
        btn_save.pack(side="right", padx=(10, 0))
        
        btn_cancel = ctk.CTkButton(btn_frame, text="Cancel", text_color="#ffffff", fg_color="gray", width=100, command=on_cancel)
        btn_cancel.pack(side="right")
        
        dialog.update_idletasks()
        width = 420
        height = 320
        parent_x = self.parent.winfo_rootx()
        parent_y = self.parent.winfo_rooty()
        parent_width = self.parent.winfo_width()
        parent_height = self.parent.winfo_height()
        x = parent_x + (parent_width // 2) - (width // 2)
        y = parent_y + (parent_height // 2) - (height // 2)
        dialog.geometry(f"{width}x{height}+{x}+{y}")
        
        ent_name.focus_set()

    def edit_ledger_item(self, tree):
        selected = tree.selection()
        if not selected:
            messagebox.showwarning("No Selection", "Please select a ledger to edit.", parent=self)
            return
            
        item_id = selected[0]
        item_vals = tree.item(item_id, 'values')
        old_name = item_vals[0]
        old_cat = item_vals[1] if len(item_vals) > 1 else ""

        dialog = ctk.CTkToplevel(self)
        dialog.title("Edit Ledger")
        dialog.transient(self)
        dialog.grab_set()
        
        main_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        ctk.CTkLabel(main_frame, text="Ledger Name:", font=("Segoe UI", 14, "bold")).pack(anchor="w", pady=(0, 5))
        ent_name = ctk.CTkEntry(main_frame, font=("Segoe UI", 16), width=360, height=40)
        ent_name.insert(0, old_name)
        ent_name.pack(fill="x", pady=(0, 15))
        
        ctk.CTkLabel(main_frame, text="Financial Report Category:", font=("Segoe UI", 14, "bold")).pack(anchor="w", pady=(0, 5))
        categories = sorted(self.parent.config_data.get("fin_report_categories", []))
        cb_category = ctk.CTkComboBox(main_frame, values=categories, state="readonly", font=("Segoe UI", 16), width=360, height=40)
        cb_category.set(old_cat)
        cb_category.pack(fill="x", pady=(0, 15))
        
        def on_cancel():
            dialog.destroy()
            
        def on_save():
            name = ent_name.get().strip()
            category = cb_category.get().strip()
            
            if not name:
                messagebox.showerror("Error", "Ledger Name is required.", parent=dialog)
                return
                
            ledgers = self.parent.config_data.get("ledgers", [])
            if name != old_name and name in ledgers:
                messagebox.showwarning("Duplicate", "This ledger already exists.", parent=dialog)
                return
                
            # Update data structures
            if name != old_name:
                idx = ledgers.index(old_name)
                ledgers[idx] = name
            
            mapping = self.parent.config_data.get("ledger_category_mapping", {})
            if old_name in mapping and name != old_name:
                del mapping[old_name]
                
            if category:
                mapping[name] = category
            else:
                if name in mapping:
                    del mapping[name]
                
            self.parent.config_data["ledgers"] = ledgers
            self.parent.config_data["ledger_category_mapping"] = mapping
            self.parent.save_config()
            self.parent.refresh_comboboxes()
            
            tree.delete(*tree.get_children())
            sorted_ledgers = sorted(ledgers)
            for idx, l in enumerate(sorted_ledgers):
                c = mapping.get(l, "")
                tag = "evenrow" if idx % 2 == 0 else "oddrow"
                tree.insert("", tk.END, values=(l, c), tags=(tag,))
            dialog.destroy()
            
        btn_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        btn_frame.pack(fill="x", pady=(10, 0))
        
        accent_color = self.parent.theme_colors.get("accent", "#0078D7")
        
        btn_save = ctk.CTkButton(btn_frame, text="OK", text_color="#ffffff", fg_color=accent_color, width=100, command=on_save)
        btn_save.pack(side="right", padx=(10, 0))
        
        btn_cancel = ctk.CTkButton(btn_frame, text="Cancel", text_color="#ffffff", fg_color="gray", width=100, command=on_cancel)
        btn_cancel.pack(side="right")
        
        dialog.update_idletasks()
        width = 420
        height = 320
        parent_x = self.parent.winfo_rootx()
        parent_y = self.parent.winfo_rooty()
        parent_width = self.parent.winfo_width()
        parent_height = self.parent.winfo_height()
        x = parent_x + (parent_width // 2) - (width // 2)
        y = parent_y + (parent_height // 2) - (height // 2)
        dialog.geometry(f"{width}x{height}+{x}+{y}")
        
        ent_name.focus_set()

    def remove_ledger_item(self, tree):
        selected = tree.selection()
        if not selected:
            messagebox.showwarning("No Selection", "Please select a ledger to remove.", parent=self)
            return
            
        item = tree.item(selected[0])
        ledger_name = item['values'][0]
        
        confirm = messagebox.askyesno("Confirm Delete", f"Are you sure you want to remove ledger '{ledger_name}'?", parent=self)
        if not confirm:
            return
            
        ledgers = self.parent.config_data.get("ledgers", [])
        if ledger_name in ledgers:
            ledgers.remove(ledger_name)
            
        mapping = self.parent.config_data.get("ledger_category_mapping", {})
        if ledger_name in mapping:
            del mapping[ledger_name]
            
        self.parent.config_data["ledgers"] = ledgers
        self.parent.config_data["ledger_category_mapping"] = mapping
        self.parent.save_config()
        self.parent.refresh_comboboxes()

        tree.delete(*tree.get_children())
        sorted_ledgers = sorted(ledgers)
        for idx, l in enumerate(sorted_ledgers):
            c = mapping.get(l, "")
            tag = "evenrow" if idx % 2 == 0 else "oddrow"
            tree.insert("", tk.END, values=(l, c), tags=(tag,))

# =========================================================================
# Phase 2: Transactions Report dialog and PDF layout classes
# =========================================================================

class TransactionsReportPDF(fpdf.FPDF):
    def __init__(self, headers, col_widths, company_info=None, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.headers = headers
        self.col_widths = col_widths
        self.company_info = company_info or {}
        
    def header(self):
        import os
        # Document Header
        
        c_name = self.company_info.get("name", "")
        c_addr = self.company_info.get("address", "")
        c_phone = self.company_info.get("phone", "")
        c_email = self.company_info.get("email", "")
        c_web = self.company_info.get("website", "")
        c_logo = self.company_info.get("logo", "")

        if c_logo and os.path.isfile(c_logo):
            try:
                # Top right corner for landscape (width=297).
                # Right edge is 282. Logo width = 30. x = 282 - 30 = 252.
                self.image(c_logo, x=252, y=5, w=30)
            except Exception:
                pass

        if c_name:
            self.set_font("Helvetica", "B", 16)
            self.set_text_color(30, 58, 138)
            self.set_x(50)
            self.multi_cell(197, 8, c_name, align="C")

            self.set_font("Helvetica", "", 10)
            self.set_text_color(75, 85, 99)
            if c_addr:
                self.set_x(50)
                self.multi_cell(197, 6, c_addr, align="C")
            if c_phone:
                self.set_x(50)
                self.multi_cell(197, 6, f"Tel: {c_phone}", align="C")
            if c_email:
                self.set_x(50)
                self.multi_cell(197, 6, f"Email: {c_email}", align="C")
            if c_web:
                self.set_x(50)
                self.multi_cell(197, 6, c_web, align="C")
            self.ln(5)

        self.set_font("Helvetica", "B", 16)
        self.set_text_color(30, 58, 138)  # Primary slate blue
        self.cell(0, 10, "Financial Records System", new_x="LMARGIN", new_y="NEXT", align="L")
        
        self.set_font("Helvetica", "B", 11)
        self.set_text_color(75, 85, 99)   # Muted gray
        self.cell(0, 8, "Transactions Report", new_x="LMARGIN", new_y="NEXT", align="L")
        
        # Decorative line
        self.set_draw_color(13, 148, 136)  # Accent Teal
        self.set_line_width(0.8)
        y_line = self.get_y()
        self.line(15, y_line, 282, y_line)
        self.ln(6)
        
        # Render Table Header
        self.set_fill_color(30, 58, 138)  # Slate blue
        self.set_text_color(255, 255, 255) # White
        self.set_draw_color(30, 58, 138)
        self.set_line_width(0.3)
        font_bold_path = "C:\\Windows\\Fonts\\arialbd.ttf"
        has_arial = False
        if os.path.exists(font_bold_path):
            try:
                import warnings
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore")
                    self.add_font("Arial", "B", font_bold_path)
                has_arial = True
            except Exception:
                pass
        line_height = 5
        max_lines = 1
        header_lines_list = []
        actual_headers = []
        
        for header_text, col_w in zip(self.headers, self.col_widths):
            self.set_font("Helvetica", "B", 9)
            if header_text.startswith("Amount") and has_arial:
                self.set_font("Arial", "B", 9)
                header_text = "Amount (\u20a6)"
            elif header_text.startswith("Amount"):
                header_text = "Amount (NGN)"
            
            actual_headers.append(header_text)
            lines = self.multi_cell(col_w - 2, line_height, header_text, dry_run=True, output="LINES")
            num_lines = len(lines) if lines else 1
            if num_lines > max_lines:
                max_lines = num_lines
            header_lines_list.append(lines)
            
        row_height = max_lines * line_height + 4 # 4mm padding
        
        orig_x = self.get_x()
        x_start = orig_x
        y_start = self.get_y()
        
        for idx, (header_text, col_w) in enumerate(zip(actual_headers, self.col_widths)):
            align = "L"
            self.set_font("Helvetica", "B", 9)
            if header_text.startswith("Amount"):
                align = "R"
                if has_arial:
                    self.set_font("Arial", "B", 9)
                    
            self.set_fill_color(30, 58, 138)
            self.rect(x_start, y_start, col_w, row_height, style="FD")
            
            lines = header_lines_list[idx]
            actual_lines = len(lines) if lines else 1
            y_offset = (row_height - (actual_lines * line_height)) / 2
            
            self.set_xy(x_start, y_start + y_offset)
            for line in lines:
                self.set_x(x_start + 1)
                self.cell(col_w - 2, line_height, line, align=align)
                self.set_y(self.get_y() + line_height)
                
            x_start += col_w
            self.set_xy(x_start, y_start)
            
        self.set_y(y_start + row_height)
        
    def footer(self):
        self.set_y(-20)
        self.set_font("Helvetica", "I", 8)
        self.set_text_color(156, 163, 175) # Light gray
        self.set_draw_color(229, 231, 235)
        self.set_line_width(0.3)
        self.line(15, self.get_y() - 2, 282, self.get_y() - 2)
        
        self.cell(0, 10, "This report contains filtered transaction entries from the system database.", align="L")
        self.set_x(262)
        self.cell(20, 10, f"Page {self.page_no()}", align="R")

    def render_row(self, keys, values, col_widths, row_idx):
        # Determine maximum height needed for the row
        line_height = 5
        max_lines = 1
        cell_lines = {}
        
        for key, val in zip(keys, values):
            col_w = col_widths[key]
            lines = self.multi_cell(col_w, line_height, str(val), dry_run=True, output="LINES")
            num_lines = len(lines) if lines else 1
            if num_lines > max_lines:
                max_lines = num_lines
            cell_lines[key] = lines
            
        row_height = max_lines * line_height + 4 # 4mm padding
        
        # Check page break
        if self.get_y() + row_height > 185:
            self.add_page()
            
        # Draw cells
        orig_x = self.get_x()
        x_start = orig_x
        y_start = self.get_y()
        
        self.set_text_color(31, 41, 55)   # #1f2937 dark grey text
        self.set_draw_color(229, 231, 235) # #e5e7eb light grey border
        self.set_line_width(0.2)
        self.set_font("Helvetica", "", 9)
        
        for key, val in zip(keys, values):
            col_w = col_widths[key]
            align = "L"
            if key == "amount":
                align = "R"
                
            # Zebra striping alternate fill
            if row_idx % 2 == 1:
                self.set_fill_color(249, 250, 251) # Zebra light gray
                self.rect(x_start, y_start, col_w, row_height, style="FD")
            else:
                self.rect(x_start, y_start, col_w, row_height, style="D")
                
            # Vertical padding & centering
            lines = cell_lines[key]
            actual_lines = len(lines) if lines else 1
            y_offset = (row_height - (actual_lines * line_height)) / 2
            
            self.set_xy(x_start, y_start + y_offset)
            for line in lines:
                self.set_x(x_start + 1.5)
                self.cell(col_w - 3, line_height, line, align=align)
                self.set_y(self.get_y() + line_height)
                
            x_start += col_w
            self.set_xy(x_start, y_start)
            
        # Move cursor below row
        self.set_xy(orig_x, y_start + row_height)

class TransactionsReportFrame(ctk.CTkFrame):
    """
    Transactions report frame that displays filtered transaction lists
    and exports them to PDF with column-visibility controls.
    """
    def __init__(self, parent_container, app_instance):
        super().__init__(parent_container, fg_color="transparent")
        self.parent = app_instance
        self.colors = self.parent.colors

        # Configure layout weights
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)

        main_frame = ctk.CTkFrame(self, fg_color="transparent")
        main_frame.grid(row=0, column=0, sticky="nsew", padx=20, pady=20)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(2, weight=1) # treeview expands

        # Create a container for the top scrollable area
        top_container = ctk.CTkFrame(main_frame, fg_color="transparent")
        top_container.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        top_container.columnconfigure(0, weight=1)

        try:
            bg_color = self.winfo_toplevel().colors["bg"]
        except AttributeError:
            bg_color = "#f8f9fa"

        top_canvas = tk.Canvas(top_container, bg=bg_color, highlightthickness=0)
        top_hsb = ttk.Scrollbar(top_container, orient="horizontal", command=top_canvas.xview)
        top_scrollable_frame = ctk.CTkFrame(top_canvas, fg_color="transparent")

        canvas_window = top_canvas.create_window((0, 0), window=top_scrollable_frame, anchor="nw")
        top_canvas.configure(xscrollcommand=top_hsb.set)

        top_canvas.grid(row=0, column=0, sticky="ew")
        top_hsb.grid(row=1, column=0, sticky="ew")

        def on_frame_configure(event):
            top_canvas.configure(scrollregion=top_canvas.bbox("all"), height=top_scrollable_frame.winfo_reqheight())
        top_scrollable_frame.bind("<Configure>", on_frame_configure)

        def on_canvas_configure(event):
            if event.width > top_scrollable_frame.winfo_reqwidth():
                top_canvas.itemconfig(canvas_window, width=event.width)
        top_canvas.bind("<Configure>", on_canvas_configure)

        # Filter panel
        filter_card = ctk.CTkFrame(top_scrollable_frame, fg_color=self.parent.theme_colors["card"], corner_radius=8, border_width=1, border_color=self.parent.theme_colors["border"])
        filter_card.grid(row=0, column=0, sticky="ew", pady=(20, 10), padx=10)

        for col_idx in range(6):
            filter_card.columnconfigure(col_idx, weight=1)

        lbl_filter_title = ttk.Label(filter_card, text="Filtering Criteria:", font=("Segoe UI", 14, "bold"), style="CardLabel.TLabel")
        lbl_filter_title.grid(row=0, column=0, columnspan=6, sticky="w", padx=10, pady=(10, 5))

        lbl_date_from = ttk.Label(filter_card, text="Date From (DD-MM-YYYY):", style="CardLabel.TLabel")
        lbl_date_from.grid(row=1, column=0, padx=10, pady=(5, 2), sticky="w")

        lbl_date_to = ttk.Label(filter_card, text="Date To (DD-MM-YYYY):", style="CardLabel.TLabel")
        lbl_date_to.grid(row=1, column=1, padx=10, pady=(5, 2), sticky="w")

        lbl_type = ttk.Label(filter_card, text="Type:", style="CardLabel.TLabel")
        lbl_type.grid(row=1, column=2, padx=10, pady=(5, 2), sticky="w")

        lbl_source = ttk.Label(filter_card, text="Source Type:", style="CardLabel.TLabel")
        lbl_source.grid(row=3, column=0, columnspan=2, padx=10, pady=(10, 2), sticky="w")

        lbl_ledger = ttk.Label(filter_card, text="Ledger Category:", style="CardLabel.TLabel")
        lbl_ledger.grid(row=3, column=2, columnspan=2, padx=10, pady=(10, 2), sticky="w")

        lbl_fin_cat = ttk.Label(filter_card, text="Financial Report Category:", style="CardLabel.TLabel")
        lbl_fin_cat.grid(row=3, column=4, columnspan=2, padx=10, pady=(10, 2), sticky="w")

        date_from_container = ctk.CTkFrame(filter_card, fg_color="transparent")
        date_from_container.grid(row=2, column=0, padx=10, pady=(0, 10), sticky="w")
        self.ent_date_from = ctk.CTkEntry(date_from_container, width=100)
        self.ent_date_from.pack(side="left")
        self.btn_date_from_picker = ctk.CTkButton(date_from_container, text="📅", width=35, text_color="#ffffff", fg_color=self.parent.theme_colors["accent"], command=lambda: self.pick_date(self.ent_date_from))
        self.btn_date_from_picker.pack(side="left", padx=(5, 0))

        date_to_container = ctk.CTkFrame(filter_card, fg_color="transparent")
        date_to_container.grid(row=2, column=1, padx=10, pady=(0, 10), sticky="w")
        self.ent_date_to = ctk.CTkEntry(date_to_container, width=100)
        self.ent_date_to.pack(side="left")
        self.btn_date_to_picker = ctk.CTkButton(date_to_container, text="📅", width=35, text_color="#ffffff", fg_color=self.parent.theme_colors["accent"], command=lambda: self.pick_date(self.ent_date_to))
        self.btn_date_to_picker.pack(side="left", padx=(5, 0))

        self.cb_type = ctk.CTkComboBox(filter_card, values=[""] + self.parent.config_data.get("types", []), width=120)
        self.cb_type.grid(row=2, column=2, padx=10, pady=(0, 10), sticky="w")
        self.cb_type.set("")

        self.cb_source_type = ctk.CTkComboBox(filter_card, values=[""] + self.parent.config_data.get("source_types", []), width=300)
        self.cb_source_type.grid(row=4, column=0, columnspan=2, padx=10, pady=(0, 10), sticky="w")
        self.cb_source_type.set("")

        self.cb_ledger = ctk.CTkComboBox(filter_card, values=[""] + self.parent.config_data.get("ledgers", []), width=300)
        self.cb_ledger.grid(row=4, column=2, columnspan=2, padx=10, pady=(0, 10), sticky="w")
        self.cb_ledger.set("")

        self.cb_fin_cat = ctk.CTkComboBox(filter_card, values=[""] + self.parent.config_data.get("fin_report_categories", []), width=300)
        self.cb_fin_cat.grid(row=4, column=4, columnspan=2, padx=10, pady=(0, 10), sticky="w")
        self.cb_fin_cat.set("")

        # Checkboxes Card for PDF columns
        cb_card = ctk.CTkFrame(top_scrollable_frame, fg_color=self.parent.theme_colors["card"], corner_radius=8, border_width=1, border_color=self.parent.theme_colors["border"])
        cb_card.grid(row=1, column=0, sticky="ew", pady=(0, 10), padx=10)

        lbl_cb = ttk.Label(cb_card, text="Include Columns in PDF Export:", font=("Segoe UI", 13, "bold"), style="CardLabel.TLabel")
        lbl_cb.pack(anchor="w", pady=(10, 5), padx=10)

        cb_row = ctk.CTkFrame(cb_card, fg_color="transparent")
        cb_row.pack(fill="x", padx=10, pady=(0, 10))

        columns_to_show = [
            ("date", "Date"),
            ("voucher", "Voucher Number"),
            ("type", "Type"),
            ("source_type", "Source Type"),
            ("source_ref", "Source Ref"),
            ("desc", "Description"),
            ("ledger", "Ledger"),
            ("fin_cat", "Financial Report Category"),
            ("amount", "Amount")
        ]

        self.col_vars = {}
        for col_key, col_title in columns_to_show:
            self.col_vars[col_key] = tk.BooleanVar(value=True)
            self.col_vars[col_key].trace_add("write", lambda *args: self.update_grid_columns())
            chk = ctk.CTkCheckBox(cb_row, text=col_title, variable=self.col_vars[col_key], text_color=self.parent.theme_colors["text"])
            chk.pack(side="left", padx=10, pady=2)

        # Grid frame
        grid_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        grid_frame.grid(row=2, column=0, sticky="nsew", pady=(0, 10))
        grid_frame.rowconfigure(0, weight=1)
        grid_frame.columnconfigure(0, weight=1)

        columns = ("No.", "Date", "Voucher Number", "Type", "Source Type", "Source Ref", "Description", "Ledger", "Financial Report Category", "Amount")
        self.tree = ttk.Treeview(grid_frame, columns=columns, show="headings", selectmode="none")

        for col in columns:
            anchor = "e" if col == "Amount" else "w"
            header_text = "Amount (₦)" if col == "Amount" else col
            self.tree.heading(col, text=header_text, anchor=anchor)
            if col == "No.":
                self.tree.column(col, width=50, minwidth=50, anchor=anchor)
            elif col == "Description":
                self.tree.column(col, width=220, minwidth=150, anchor=anchor)
            elif col in ("Date", "Type", "Amount"):
                self.tree.column(col, width=90, minwidth=80, anchor=anchor)
            else:
                self.tree.column(col, width=110, minwidth=100, anchor=anchor)

        self.tree.grid(row=0, column=0, sticky="nsew")

        theme = self.parent.config_data.get("theme", "light")
        even_bg = "#121212" if theme == "dark" else "#f9fafb"
        odd_bg  = "#000000" if theme == "dark" else "#ffffff"
        text_fg = "#f9fafb" if theme == "dark" else "#1f2937"
        self.tree.tag_configure('evenrow', background=even_bg, foreground=text_fg)
        self.tree.tag_configure('oddrow', background=odd_bg, foreground=text_fg)

        v_scroll = ttk.Scrollbar(grid_frame, orient="vertical", command=self.tree.yview)
        v_scroll.grid(row=0, column=1, sticky="ns")
        h_scroll = ttk.Scrollbar(grid_frame, orient="horizontal", command=self.tree.xview)
        h_scroll.grid(row=1, column=0, sticky="ew")
        self.tree.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)

        self.lbl_record_count = ttk.Label(grid_frame, text="Total Records: 0", font=("Segoe UI", 11, "bold"), foreground="#64748B")
        self.lbl_record_count.grid(row=2, column=0, columnspan=2, sticky="w", pady=(5, 0))

        # Bottom toolbar
        bottom_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        bottom_frame.grid(row=3, column=0, sticky="ew")

        self.btn_print_report = ctk.CTkButton(bottom_frame, text="Print Report", text_color="#ffffff", fg_color=self.parent.theme_colors["accent"], command=self.print_to_pdf)
        self.btn_print_report.pack(side="left")

        self.btn_apply = ctk.CTkButton(bottom_frame, text="Apply Filters", text_color="#ffffff", fg_color=self.parent.theme_colors["accent"], command=self.apply_filters)
        self.btn_apply.pack(side="right", padx=(5, 0))

        self.btn_clear_filters = ctk.CTkButton(bottom_frame, text="Clear Filters", text_color="#ffffff", fg_color=self.parent.theme_colors["accent"], command=self.clear_filters)
        self.btn_clear_filters.pack(side="right", padx=5)
        
        # Load initial values
        self.load_data()

    def pick_date(self, entry_widget):
        """Opens a popup calendar to select a date."""
        CalendarDialog(self, entry_widget)

    def update_grid_columns(self):
        """Updates treeview displaycolumns based on checkboxes."""
        display_cols = []
        col_mapping = {
            "date": "Date",
            "voucher": "Voucher Number",
            "type": "Type",
            "source_type": "Source Type",
            "source_ref": "Source Ref",
            "desc": "Description",
            "ledger": "Ledger",
            "fin_cat": "Financial Report Category",
            "amount": "Amount"
        }
        for col_key, col_title in col_mapping.items():
            if self.col_vars[col_key].get():
                display_cols.append(col_title)
        self.tree.configure(displaycolumns=display_cols)

    def load_data(self):
        """Loads data from the active workbook into self.all_records and populates the treeview."""
        self.all_records = []
        if not hasattr(self.parent, 'current_workbook_path') or not self.parent.current_workbook_path:
            self.apply_filters()
            return
            
        try:
            wb = openpyxl.load_workbook(self.parent.current_workbook_path, data_only=True)
            ws_trans = wb["TransactionEntries"]
            
            for row_idx in range(2, ws_trans.max_row + 1):
                row_vals = [ws_trans.cell(row=row_idx, column=c).value for c in range(1, 10)]
                if all(v is None for v in row_vals):
                    continue
                
                # Format Date
                if isinstance(row_vals[0], datetime):
                    date_str = row_vals[0].strftime("%d/%m/%Y")
                elif row_vals[0] is not None:
                    date_str = str(row_vals[0]).strip()
                    try:
                        dt = datetime.strptime(date_str, "%Y-%m-%d")
                        date_str = dt.strftime("%d/%m/%Y")
                    except ValueError:
                        pass
                else:
                    date_str = ""
                    
                # Format Amount
                amt_str = ""
                if row_vals[8] is not None:
                    try:
                        amt_str = f"{float(row_vals[8]):,.2f}"
                    except ValueError:
                        amt_str = str(row_vals[8])
                        
                self.all_records.append({
                    "row_idx": row_idx,
                    "date": date_str,
                    "voucher": str(row_vals[1] if row_vals[1] is not None else "").strip(),
                    "type": str(row_vals[2] if row_vals[2] is not None else "").strip(),
                    "source_type": str(row_vals[3] if row_vals[3] is not None else "").strip(),
                    "source_ref": str(row_vals[4] if row_vals[4] is not None else "").strip(),
                    "desc": str(row_vals[5] if row_vals[5] is not None else "").strip(),
                    "ledger": str(row_vals[6] if row_vals[6] is not None else "").strip(),
                    "fin_cat": str(row_vals[7] if row_vals[7] is not None else "").strip(),
                    "amount": amt_str
                })
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load workbook data:\n{e}", parent=self)
            
        self.apply_filters()

    def apply_filters(self):
        """Applies current filter selections to the list of records and populates treeview."""
        for item in self.tree.get_children():
            self.tree.delete(item)
            
        date_from_str = self.ent_date_from.get().strip()
        date_to_str = self.ent_date_to.get().strip()
        type_val = self.cb_type.get().strip()
        source_val = self.cb_source_type.get().strip()
        ledger_val = self.cb_ledger.get().strip()
        fin_cat_val = self.cb_fin_cat.get().strip()
        
        date_from = None
        if date_from_str:
            try:
                date_from = datetime.strptime(date_from_str, "%d/%m/%Y").date()
            except ValueError:
                try:
                    date_from = datetime.strptime(date_from_str, "%d-%m-%Y").date()
                except ValueError:
                    try:
                        date_from = datetime.strptime(date_from_str, "%Y-%m-%d").date()
                    except ValueError:
                        messagebox.showerror("Validation Error", "Date From must be in DD-MM-YYYY format.", parent=self)
                        return
                
        date_to = None
        if date_to_str:
            try:
                date_to = datetime.strptime(date_to_str, "%d/%m/%Y").date()
            except ValueError:
                try:
                    date_to = datetime.strptime(date_to_str, "%d-%m-%Y").date()
                except ValueError:
                    try:
                        date_to = datetime.strptime(date_to_str, "%Y-%m-%d").date()
                    except ValueError:
                        messagebox.showerror("Validation Error", "Date To must be in DD-MM-YYYY format.", parent=self)
                        return
        
        if date_from and date_to and date_from > date_to:
            messagebox.showerror("Validation Error", "'Date From' cannot be later than 'Date To'.", parent=self)
            return
                
        self.filtered_records = []
        for rec in self.all_records:
            rec_date_str = rec["date"]
            rec_date = None
            if rec_date_str:
                try:
                    rec_date = datetime.strptime(rec_date_str, "%d/%m/%Y").date()
                except ValueError:
                    try:
                        rec_date = datetime.strptime(rec_date_str, "%Y-%m-%d").date()
                    except ValueError:
                        pass
            
            # Match rules
            if date_from and (not rec_date or rec_date < date_from):
                continue
            if date_to and (not rec_date or rec_date > date_to):
                continue
            if type_val and rec["type"] != type_val:
                continue
            if source_val and rec["source_type"] != source_val:
                continue
            if ledger_val and rec["ledger"] != ledger_val:
                continue
            if fin_cat_val and rec["fin_cat"] != fin_cat_val:
                continue
                
            self.filtered_records.append(rec)
            
        for idx, rec in enumerate(self.filtered_records, 1):
            tag = "evenrow" if idx % 2 == 0 else "oddrow"
            self.tree.insert(
                "",
                tk.END,
                values=(
                    idx,
                    rec["date"],
                    rec["voucher"],
                    rec["type"],
                    rec["source_type"],
                    rec["source_ref"],
                    rec["desc"],
                    rec["ledger"],
                    rec["fin_cat"],
                    rec["amount"]
                ),
                tags=(tag,)
            )

        if hasattr(self, "lbl_record_count"):
            self.lbl_record_count.configure(text=f"Total Records: {len(self.filtered_records)}")

    def clear_filters(self):
        """Clears all filters and displays all records."""
        self.ent_date_from.delete(0, tk.END)
        self.ent_date_to.delete(0, tk.END)
        self.cb_type.set("")
        self.cb_source_type.set("")
        self.cb_ledger.set("")
        self.cb_fin_cat.set("")
        self.apply_filters()

    def print_to_pdf(self):
        """Generates a clean tabular PDF of the currently filtered transaction entries."""
        if not self.filtered_records:
            messagebox.showwarning("No Data", "There are no transactions to print.", parent=self)
            return

        if fpdf is None:
            messagebox.showerror(
                "Error",
                "The 'fpdf2' library is required to export reports.\n"
                "Please run: pip install fpdf2",
                parent=self
            )
            return

        dirs = self.parent.config_data.get("directories", {})
        reports_dir = dirs.get("transactions_report_directory", "").strip()
        
        if not reports_dir or not os.path.isdir(reports_dir):
            initial_dir = os.path.dirname(self.parent.current_workbook_path)
            reports_dir = os.path.join(initial_dir, "reports")
            
        os.makedirs(reports_dir, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        pdf_path = os.path.join(reports_dir, f"Transactions_Report_{timestamp}.pdf")
            
        pdf_filename = os.path.basename(pdf_path)

        # Determine selected columns
        selected_cols = [("sn", "S/N")]
        for col_key, col_title in [
            ("date", "Date"),
            ("voucher", "Voucher No"),
            ("type", "Type"),
            ("source_type", "Source Type"),
            ("source_ref", "Source Ref"),
            ("desc", "Description"),
            ("ledger", "Ledger"),
            ("fin_cat", "Financial Category"),
            ("amount", "Amount")
        ]:
            if self.col_vars[col_key].get():
                selected_cols.append((col_key, col_title))
                
        if not selected_cols:
            messagebox.showwarning("No Columns", "Please select at least one column to include in the PDF.", parent=self)
            return

        col_weights = {
            "sn": 0.8,
            "date": 2.0,
            "voucher": 2.5,
            "type": 2.0,
            "source_type": 2.5,
            "source_ref": 2.5,
            "desc": 5.0,
            "ledger": 2.5,
            "fin_cat": 2.5,
            "amount": 2.5
        }

        # Printable width 267mm
        printable_width = 267.0
        total_weight = sum(col_weights[col_key] for col_key, _ in selected_cols)
        
        widths = {}
        for col_key, _ in selected_cols:
            widths[col_key] = (col_weights[col_key] / total_weight) * printable_width

        headers = [col_title for _, col_title in selected_cols]
        col_widths_list = [widths[col_key] for col_key, _ in selected_cols]

        try:
            company_info = self.parent.config_data.get("company_info", {})
            pdf = TransactionsReportPDF(headers, col_widths_list, company_info=company_info, orientation="L", unit="mm", format="A4")
            pdf.set_margins(15, 15, 15)
            pdf.add_page()
            
            all_col_titles = ["Date", "Voucher No", "Type", "Source Type", "Source Ref", "Description", "Ledger", "Financial Category", "Amount"]
            selected_titles = [t for _, t in selected_cols]
            unselected_cols = [t for t in all_col_titles if t not in selected_titles]

            active_filters = {}
            if self.ent_date_from.get().strip() or self.ent_date_to.get().strip():
                active_filters["Date Range"] = f"{self.ent_date_from.get().strip()} to {self.ent_date_to.get().strip()}"
            if self.cb_type.get().strip():
                active_filters["Type"] = self.cb_type.get().strip()
            if self.cb_source_type.get().strip():
                active_filters["Source Type"] = self.cb_source_type.get().strip()
            if self.cb_ledger.get().strip():
                active_filters["Ledger Category"] = self.cb_ledger.get().strip()
            if self.cb_fin_cat.get().strip():
                active_filters["Fin Category"] = self.cb_fin_cat.get().strip()
            active_filters_str = ", ".join(f"{k}: {v}" for k, v in active_filters.items())


            row_idx = 1
            for rec in self.filtered_records:
                row_vals = []
                row_keys = []
                for col_key, _ in selected_cols:
                    if col_key == "sn":
                        val = str(row_idx)
                    else:
                        val = rec[col_key]
                        if col_key == "amount":
                            try:
                                val = f"{float(val.replace(',', '')):,.2f}"
                            except (ValueError, TypeError):
                                pass
                    row_vals.append(val)
                    row_keys.append(col_key)
                    
                pdf.render_row(row_keys, row_vals, widths, row_idx)
                row_idx += 1

            pdf.ln(5)
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_text_color(31, 41, 55)
            status_text = f"Total Records: {len(self.filtered_records)}"
            pdf.cell(0, 8, status_text, new_x="LMARGIN", new_y="NEXT", align="L")
            
            if unselected_cols:
                pdf.set_font("Helvetica", "I", 9)
                pdf.set_text_color(75, 85, 99)
                pdf.cell(0, 8, "Unselected Columns: " + ", ".join(unselected_cols), new_x="LMARGIN", new_y="NEXT", align="L")
                
            if active_filters_str:
                pdf.set_font("Helvetica", "I", 9)
                pdf.set_text_color(75, 85, 99)
                pdf.cell(0, 8, "Filters Applied: " + active_filters_str, new_x="LMARGIN", new_y="NEXT", align="L")
                
            pdf.output(pdf_path)

        except PermissionError:
            messagebox.showerror(
                "Permission Error",
                f"Could not save the PDF file because it is currently open in another application.\n"
                f"Please close the file '{pdf_filename}' and try again.",
                parent=self
            )
            return
        except Exception as e:
            messagebox.showerror(
                "Error",
                f"An error occurred while exporting the report:\n{e}",
                parent=self
            )
            return

        try:
            os.startfile(pdf_path)
        except Exception as e:
            messagebox.showerror(
                "Error",
                f"Failed to open the generated PDF report:\n{e}",
                parent=self
            )

    # on_close method removed since it is now a frame

class CalendarDialog(ctk.CTkToplevel):
    def __init__(self, parent, target_entry):
        super().__init__(parent)
        self.parent = parent
        self.target_entry = target_entry
        self.title("Select Date")
        self.geometry("280x300")
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()
        
        self.center_window()
        
        if hasattr(parent, "parent") and hasattr(parent.parent, "theme_colors"):
            self.colors = parent.parent.theme_colors
        elif hasattr(parent, "theme_colors"):
            self.colors = parent.theme_colors
        else:
            self.colors = {"bg": "#ffffff", "card": "#ffffff", "text": "#000000", "accent": "#1f6aa5"}
            
        self.configure(fg_color=self.colors["bg"])
        
        self.now = datetime.now()
        self.current_year = self.now.year
        self.current_month = self.now.month
        
        entry_val = target_entry.get().strip()
        if entry_val:
            try:
                dt = datetime.strptime(entry_val, "%d-%m-%Y")
                self.current_year = dt.year
                self.current_month = dt.month
            except ValueError:
                pass
                
        header = ctk.CTkFrame(self, fg_color="transparent")
        header.pack(fill="x", pady=(10, 5), padx=10)
        
        btn_prev_year = ctk.CTkButton(header, text="<<", width=30, command=self.prev_year)
        btn_prev_year.pack(side="left", padx=(0, 2))
        
        btn_prev = ctk.CTkButton(header, text="<", width=30, command=self.prev_month)
        btn_prev.pack(side="left")
        
        self.lbl_month = ctk.CTkLabel(header, text="", font=("Segoe UI", 13, "bold"))
        self.lbl_month.pack(side="left", expand=True)
        
        btn_next_year = ctk.CTkButton(header, text=">>", width=30, command=self.next_year)
        btn_next_year.pack(side="right")
        
        btn_next = ctk.CTkButton(header, text=">", width=30, command=self.next_month)
        btn_next.pack(side="right", padx=(0, 2))
        
        self.days_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.days_frame.pack(fill="both", expand=True, padx=10, pady=(5, 10))
        
        self.draw_calendar()
        
    def center_window(self):
        self.update_idletasks()
        width = 280
        height = 300
        main_win = self.parent.winfo_toplevel()
        main_x = main_win.winfo_rootx()
        main_y = main_win.winfo_rooty()
        main_w = main_win.winfo_width()
        main_h = main_win.winfo_height()
        x = main_x + (main_w - width) // 2
        y = main_y + (main_h - height) // 2
        self.geometry(f"{width}x{height}+{x}+{y}")
        
    def prev_month(self):
        self.current_month -= 1
        if self.current_month < 1:
            self.current_month = 12
            self.current_year -= 1
        self.draw_calendar()
        
    def next_month(self):
        self.current_month += 1
        if self.current_month > 12:
            self.current_month = 1
            self.current_year += 1
        self.draw_calendar()
        
    def prev_year(self):
        self.current_year -= 1
        self.draw_calendar()
        
    def next_year(self):
        self.current_year += 1
        self.draw_calendar()
        
    def draw_calendar(self):
        import calendar
        for child in self.days_frame.winfo_children():
            child.destroy()
            
        month_name = calendar.month_name[self.current_month]
        self.lbl_month.configure(text=f"{month_name} {self.current_year}")
        
        headers = ["Mo", "Tu", "We", "Th", "Fr", "Sa", "Su"]
        for idx, h in enumerate(headers):
            lbl = ctk.CTkLabel(self.days_frame, text=h, font=("Segoe UI", 11, "bold"))
            lbl.grid(row=0, column=idx, sticky="ew")
            
        cal = calendar.monthcalendar(self.current_year, self.current_month)
        accent = self.colors.get("accent", "#1f6aa5")
        for r_idx, week in enumerate(cal, 1):
            for c_idx, day in enumerate(week):
                if day == 0:
                    continue
                btn = ctk.CTkButton(
                    self.days_frame,
                    text=str(day),
                    width=30,
                    height=30,
                    font=("Segoe UI", 11),
                    fg_color="transparent",
                    text_color=self.colors["text"],
                    hover_color=accent,
                    command=lambda d=day: self.select_day(d)
                )
                btn.grid(row=r_idx, column=c_idx, sticky="nsew", padx=1, pady=1)
                
        for i in range(7):
            self.days_frame.columnconfigure(i, weight=1)
        for i in range(7):
            self.days_frame.rowconfigure(i, weight=1)
            
    def select_day(self, day):
        date_str = f"{day:02d}-{self.current_month:02d}-{self.current_year:04d}"
        orig_state = self.target_entry.cget("state")
        self.target_entry.configure(state="normal")
        self.target_entry.delete(0, tk.END)
        self.target_entry.insert(0, date_str)
        try:
            self.target_entry.configure(state=orig_state)
        except Exception:
            pass
        self.grab_release()
        self.destroy()

class MultilineEntry(tk.Text):
    def __init__(self, master, app, **kwargs):
        self.app = app
        border_color = app.colors["border"] if hasattr(app, "colors") else "#e5e7eb"
        bg_color = app.colors["card"] if hasattr(app, "colors") else "#ffffff"
        fg_color = app.colors["text"] if hasattr(app, "colors") else "#1f2937"
        
        kwargs.setdefault("bd", 1)
        kwargs.setdefault("relief", "solid")
        kwargs.setdefault("highlightthickness", 0)
        kwargs.setdefault("bg", bg_color)
        kwargs.setdefault("fg", fg_color)
        kwargs.setdefault("font", ("Segoe UI", 10))
        
        super().__init__(master, **kwargs)
        
    def get(self, *args):
        if not args:
            return super().get("1.0", "end-1c")
        return super().get(*args)
        
    def delete(self, start, end=None):
        if start == 0 and (end == tk.END or end == "end"):
            super().delete("1.0", tk.END)
        else:
            super().delete(start, end)
            
    def insert(self, index, text):
        if index == 0 or index == "0":
            super().insert("1.0", text)
        else:
            super().insert(index, text)
            
    def configure(self, cnf=None, **kw):
        if "state" in kw:
            state = kw["state"]
            if state == "disabled":
                kw["bg"] = self.app.colors["bg"]
            else:
                kw["bg"] = self.app.colors["card"]
        super().configure(cnf, **kw)
        
    def config(self, cnf=None, **kw):
        self.configure(cnf, **kw)

if __name__ == "__main__":
    app = App()
    app.mainloop()
